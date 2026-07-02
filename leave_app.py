"""
Leave Application Module — integrates with existing attendance_data.xlsx
Run standalone:  python leave_app.py  →  http://localhost:5001
Or import into app.py and register the blueprint.

Two roles:
  • Student  – login with Student ID + password (default = student_id)
              Can: apply for leave, view OWN leaves only
  • Admin    – login with username "admin" + password "admin123"
              Can: view ALL leave applications, approve / reject with remarks
"""

from flask import (Flask, render_template_string, request,
                   jsonify, session, redirect, url_for)
from datetime import datetime
import openpyxl
from openpyxl.styles import Font, PatternFill, Alignment, Border, Side
from openpyxl.utils import get_column_letter
import os, json, hashlib, secrets, re, html, time
from collections import defaultdict
from werkzeug.security import check_password_hash, generate_password_hash

# ─── CONFIG ───────────────────────────────────────────────────
EXCEL_FILE   = "attendance_data.xlsx"
LEAVE_FILE   = "leave_data.xlsx"
ADMIN_USER   = "admin"
ADMIN_PASS   = "admin123"   # change in production
ADMIN_PASS_HASH = generate_password_hash(ADMIN_PASS)
SECRET_KEY   = secrets.token_hex(16)

app = Flask(__name__)
app.secret_key = SECRET_KEY

# Secure Session Cookies
app.config.update(
    SESSION_COOKIE_HTTPONLY=True,
    SESSION_COOKIE_SECURE=True,
    SESSION_COOKIE_SAMESITE='Lax'
)

# ─────────────────────────────────────────────────────────────
#  SECURITY HELPERS, RATE LIMITING & CSRF PROTECTION
# ─────────────────────────────────────────────────────────────

def is_valid_email(email):
    regex = r'^[a-zA-Z0-9._%+-]+@[a-zA-Z0-9.-]+\.[a-zA-Z]{2,}$'
    return bool(re.match(regex, email))

def is_valid_date(date_str):
    try:
        datetime.strptime(date_str, "%Y-%m-%d")
        return True
    except ValueError:
        return False

def is_alphanumeric(val):
    return bool(re.match(r'^[a-zA-Z0-9_-]+$', val))

class RateLimiter:
    def __init__(self, limit=5, window=60):
        self.limit = limit
        self.window = window
        self.attempts = defaultdict(list)
        
    def is_blocked(self, key):
        now = time.time()
        self.attempts[key] = [t for t in self.attempts[key] if now - t < self.window]
        if len(self.attempts[key]) >= self.limit:
            return True
        self.attempts[key].append(now)
        return False

login_limiter = RateLimiter(limit=5, window=60)

@app.before_request
def security_before_request():
    if 'csrf_token' not in session:
        session['csrf_token'] = secrets.token_hex(32)
        
    if request.method in ["POST", "PUT", "DELETE", "PATCH"]:
        csrf_token = request.headers.get("X-CSRFToken")
        if not csrf_token or csrf_token != session.get("csrf_token"):
            return jsonify(success=False, error="CSRF", message="CSRF token missing or invalid."), 400

@app.after_request
def add_security_headers(response):
    response.headers['X-Frame-Options'] = 'SAMEORIGIN'
    response.headers['X-Content-Type-Options'] = 'nosniff'
    response.headers['X-XSS-Protection'] = '1; mode=block'
    response.headers['Referrer-Policy'] = 'strict-origin-when-cross-origin'
    
    # CSP: Allow self, Google Fonts, and chart.js via jsdelivr CDN
    csp = (
        "default-src 'self'; "
        "script-src 'self' 'unsafe-inline' https://cdn.jsdelivr.net; "
        "style-src 'self' 'unsafe-inline' https://fonts.googleapis.com; "
        "font-src 'self' https://fonts.gstatic.com; "
        "img-src 'self' data:; "
        "connect-src 'self';"
    )
    response.headers['Content-Security-Policy'] = csp
    return response

# ─── EXCEL HELPERS ────────────────────────────────────────────
def thin():
    s = Side(style="thin", color="D1D5DB")
    return Border(left=s, right=s, top=s, bottom=s)

def hdr(cell, bg="1E3A5F"):
    cell.fill      = PatternFill("solid", fgColor=bg)
    cell.font      = Font(bold=True, color="FFFFFF", name="Calibri", size=10)
    cell.alignment = Alignment(horizontal="center", vertical="center")
    cell.border    = thin()

def cell_style(cell, bg="FFFFFF"):
    cell.fill      = PatternFill("solid", fgColor=bg)
    cell.font      = Font(name="Calibri", size=10)
    cell.alignment = Alignment(horizontal="center", vertical="center")
    cell.border    = thin()

LEAVE_STATUS_BG = {
    "Pending":  "FFFBEB",
    "Approved": "F0FDF4",
    "Rejected": "FEF2F2",
}

def get_leave_wb():
    if os.path.exists(LEAVE_FILE):
        return openpyxl.load_workbook(LEAVE_FILE)
    wb = openpyxl.Workbook()
    wb.remove(wb.active)
    ws = wb.create_sheet("LeaveApplications")
    cols = ["Leave ID","Student ID","Student Name","Class",
            "Leave Type","From Date","To Date","Days","Reason",
            "Status","Admin Remarks","Applied On","Reviewed On"]
    ws.row_dimensions[1].height = 28
    for i, h in enumerate(cols, 1):
        hdr(ws.cell(1, i, h))
    widths = [12,12,22,12,15,13,13,8,35,12,30,20,20]
    for i, w in enumerate(widths, 1):
        ws.column_dimensions[get_column_letter(i)].width = w
    wb.save(LEAVE_FILE)
    return wb

def get_students_list():
    """Pull students from the main attendance workbook."""
    if not os.path.exists(EXCEL_FILE):
        return {}
    try:
        wb = openpyxl.load_workbook(EXCEL_FILE)
        ws = wb["Students"]
        out = {}
        for row in ws.iter_rows(min_row=2, values_only=True):
            if row[0]:
                sid = str(row[0]).strip()
                out[sid] = {
                    "name":  str(row[1]).strip() if row[1] else "",
                    "class": str(row[2]).strip() if row[2] else "",
                    "email": str(row[3]).strip() if row[3] else "",
                }
        return out
    except Exception:
        return {}

def next_leave_id(ws):
    max_id = 0
    for row in ws.iter_rows(min_row=2, values_only=True):
        if row[0]:
            try:
                num = int(str(row[0]).replace("LV",""))
                if num > max_id:
                    max_id = num
            except:
                pass
    return f"LV{max_id+1:04d}"

def date_diff(from_d, to_d):
    try:
        f = datetime.strptime(from_d, "%Y-%m-%d")
        t = datetime.strptime(to_d,   "%Y-%m-%d")
        return max(1, (t - f).days + 1)
    except:
        return 1

# ─── AUTH HELPERS ─────────────────────────────────────────────
def is_admin():
    return session.get("role") == "admin"

def is_student():
    return session.get("role") == "student"

def current_student_id():
    return session.get("student_id")

# ─── ROUTES ───────────────────────────────────────────────────

@app.route("/leave")
def leave_index():
    return render_template_string(HTML_PAGE)

@app.route("/leave/api/auth/login", methods=["POST"])
def auth_login():
    ip = request.remote_addr
    if login_limiter.is_blocked(ip):
        return jsonify(success=False, message="Too many login attempts. Please try again in a minute."), 429

    d = request.get_json() or {}
    username = d.get("username", "").strip()
    password = d.get("password", "").strip()

    if not username or not password:
        return jsonify(success=False, message="Username and Password are required."), 400
    if len(username) > 100 or len(password) > 100:
        return jsonify(success=False, message="Inputs exceed maximum allowed length."), 400

    # Admin login
    if username == ADMIN_USER and check_password_hash(ADMIN_PASS_HASH, password):
        csrf = session.get("csrf_token")
        session.clear()
        if csrf:
            session["csrf_token"] = csrf
        else:
            session["csrf_token"] = secrets.token_hex(32)
        session["role"]     = "admin"
        session["username"] = ADMIN_USER
        return jsonify(success=True, role="admin", name="Administrator")

    # Student login  — username = student_id, password = student_id (default)
    students = get_students_list()
    if username in students:
        stored_pass = _get_student_password(username)
        password_correct = False
        if stored_pass:
            if stored_pass.startswith(('pbkdf2:', 'scrypt:', 'bcrypt:')):
                password_correct = check_password_hash(stored_pass, password)
            else:
                password_correct = (stored_pass == password)
        else:
            password_correct = (password == username)

        if password_correct:
            csrf = session.get("csrf_token")
            session.clear()
            if csrf:
                session["csrf_token"] = csrf
            else:
                session["csrf_token"] = secrets.token_hex(32)
            session["role"]       = "student"
            session["student_id"] = username
            session["username"]   = username
            return jsonify(success=True, role="student",
                           name=students[username]["name"],
                           student_id=username)

    return jsonify(success=False, message="Invalid credentials.")

def _get_student_password(sid):
    """Simple password store in leave_data.xlsx – Passwords sheet."""
    if not os.path.exists(LEAVE_FILE):
        return None
    wb = openpyxl.load_workbook(LEAVE_FILE)
    if "Passwords" not in wb.sheetnames:
        return None
    ws = wb["Passwords"]
    for row in ws.iter_rows(min_row=2, values_only=True):
        if row[0] and str(row[0]).strip() == str(sid):
            return str(row[1]).strip() if row[1] else None
    return None

@app.route("/leave/api/auth/logout", methods=["POST"])
def auth_logout():
    csrf = session.get("csrf_token")
    session.clear()
    if csrf:
        session["csrf_token"] = csrf
    else:
        session["csrf_token"] = secrets.token_hex(32)
    return jsonify(success=True)

@app.route("/leave/api/auth/me")
def auth_me():
    if is_admin():
        return jsonify(logged_in=True, role="admin", name="Administrator")
    if is_student():
        students = get_students_list()
        sid = current_student_id()
        name = students.get(sid, {}).get("name", sid)
        return jsonify(logged_in=True, role="student", name=name, student_id=sid)
    return jsonify(logged_in=False)

# ── Apply for leave (student only) ───
@app.route("/leave/api/apply", methods=["POST"])
def apply_leave():
    if not is_student():
        return jsonify(success=False, message="Unauthorised."), 401

    d = request.get_json() or {}
    sid        = current_student_id()
    leave_type = d.get("leave_type","").strip()
    from_date  = d.get("from_date","").strip()
    to_date    = d.get("to_date","").strip()
    reason     = d.get("reason","").strip()

    if not all([leave_type, from_date, to_date, reason]):
        return jsonify(success=False, message="All fields are required."), 400
    if len(leave_type) > 100 or len(from_date) > 20 or len(to_date) > 20 or len(reason) > 1000:
        return jsonify(success=False, message="Inputs exceed maximum allowed length."), 400
    if not is_valid_date(from_date) or not is_valid_date(to_date):
        return jsonify(success=False, message="Invalid date format. Expected YYYY-MM-DD."), 400

    # Escape inputs
    leave_type = html.escape(leave_type)
    reason = html.escape(reason)

    students = get_students_list()
    student  = students.get(sid)
    if not student:
        return jsonify(success=False, message="Student not found.")

    days      = date_diff(from_date, to_date)
    applied   = datetime.now().strftime("%Y-%m-%d %H:%M:%S")

    wb = get_leave_wb()
    ws = wb["LeaveApplications"]
    lid = next_leave_id(ws)
    nr  = ws.max_row + 1

    vals = [lid, sid, student["name"], student["class"],
            leave_type, from_date, to_date, days,
            reason, "Pending", "—", applied, "—"]
    for i, v in enumerate(vals, 1):
        c = ws.cell(nr, i, v)
        cell_style(c, LEAVE_STATUS_BG["Pending"])
    wb.save(LEAVE_FILE)

    return jsonify(success=True,
                   message=f"Leave application {lid} submitted successfully.",
                   leave_id=lid)

# ── Student: view MY leaves ───
@app.route("/leave/api/my_leaves")
def my_leaves():
    if not is_student():
        return jsonify(success=False, message="Unauthorised."), 401

    sid = current_student_id()
    wb  = get_leave_wb()
    ws  = wb["LeaveApplications"]
    out = []
    for row in ws.iter_rows(min_row=2, values_only=True):
        if row[1] and str(row[1]).strip() == sid:
            out.append(_row_to_dict(row))
    # Most recent first
    out.sort(key=lambda x: x["applied_on"], reverse=True)
    return jsonify(success=True, leaves=out)

# ── Admin: view ALL leaves ───
@app.route("/leave/api/all_leaves")
def all_leaves():
    if not is_admin():
        return jsonify(success=False, message="Unauthorised."), 401

    wb  = get_leave_wb()
    ws  = wb["LeaveApplications"]
    out = []
    for row in ws.iter_rows(min_row=2, values_only=True):
        if row[0]:
            out.append(_row_to_dict(row))
    out.sort(key=lambda x: x["applied_on"], reverse=True)

    # Stats
    total    = len(out)
    pending  = sum(1 for r in out if r["status"] == "Pending")
    approved = sum(1 for r in out if r["status"] == "Approved")
    rejected = sum(1 for r in out if r["status"] == "Rejected")

    return jsonify(success=True, leaves=out,
                   stats=dict(total=total, pending=pending,
                              approved=approved, rejected=rejected))

# ── Admin: approve / reject ───
@app.route("/leave/api/review", methods=["POST"])
def review_leave():
    if not is_admin():
        return jsonify(success=False, message="Unauthorised."), 401

    d       = request.get_json() or {}
    lid     = d.get("leave_id","").strip()
    action  = d.get("action","").strip()       # "Approved" | "Rejected"
    remarks = d.get("remarks","").strip()

    if not lid or not action:
        return jsonify(success=False, message="Leave ID and Action are required."), 400
    if action not in ("Approved","Rejected"):
        return jsonify(success=False, message="Invalid action."), 400
    if len(lid) > 50 or len(remarks) > 1000:
        return jsonify(success=False, message="Inputs exceed maximum allowed length."), 400
    if not is_alphanumeric(lid.replace("LV", "")):
        return jsonify(success=False, message="Invalid Leave ID format."), 400

    # Escape inputs
    lid = html.escape(lid)
    remarks = html.escape(remarks)

    wb = get_leave_wb()
    ws = wb["LeaveApplications"]
    reviewed = datetime.now().strftime("%Y-%m-%d %H:%M:%S")

    for row in ws.iter_rows(min_row=2):
        if row[0].value and str(row[0].value).strip() == lid:
            if str(row[9].value).strip() != "Pending":
                return jsonify(success=False,
                               message="This application has already been reviewed.")
            row[9].value  = action
            row[10].value = remarks or "—"
            row[12].value = reviewed
            bg = LEAVE_STATUS_BG[action]
            for cell in row:
                cell_style(cell, bg)
            wb.save(LEAVE_FILE)
            return jsonify(success=True,
                           message=f"Leave {lid} has been {action.lower()}.")

    return jsonify(success=False, message="Leave application not found.")

# ── Admin: leave stats per student ───
@app.route("/leave/api/student_summary")
def student_summary():
    if not is_admin():
        return jsonify(success=False, message="Unauthorised."), 401

    wb = get_leave_wb()
    ws = wb["LeaveApplications"]
    summary = {}
    for row in ws.iter_rows(min_row=2, values_only=True):
        if not row[0]: continue
        sid  = str(row[1]).strip()
        name = str(row[2]).strip()
        st   = str(row[9]).strip()
        if sid not in summary:
            summary[sid] = {"student_id":sid,"name":name,
                            "total":0,"pending":0,"approved":0,"rejected":0,"total_days":0}
        summary[sid]["total"] += 1
        if st == "Pending":  summary[sid]["pending"]  += 1
        elif st == "Approved":
            summary[sid]["approved"] += 1
            try: summary[sid]["total_days"] += int(row[7]) if row[7] else 0
            except: pass
        elif st == "Rejected": summary[sid]["rejected"] += 1
    return jsonify(success=True, summary=list(summary.values()))

def _row_to_dict(row):
    return {
        "leave_id":     str(row[0]).strip() if row[0] else "",
        "student_id":   str(row[1]).strip() if row[1] else "",
        "student_name": str(row[2]).strip() if row[2] else "",
        "class":        str(row[3]).strip() if row[3] else "",
        "leave_type":   str(row[4]).strip() if row[4] else "",
        "from_date":    str(row[5]).strip() if row[5] else "",
        "to_date":      str(row[6]).strip() if row[6] else "",
        "days":         str(row[7]).strip() if row[7] else "",
        "reason":       str(row[8]).strip() if row[8] else "",
        "status":       str(row[9]).strip() if row[9] else "",
        "remarks":      str(row[10]).strip() if row[10] else "—",
        "applied_on":   str(row[11]).strip() if row[11] else "",
        "reviewed_on":  str(row[12]).strip() if row[12] else "—",
    }

# ─── FRONTEND HTML ────────────────────────────────────────────
HTML_PAGE = """<!DOCTYPE html>
<html lang="en">
<head>
<meta charset="UTF-8"/>
<meta name="viewport" content="width=device-width, initial-scale=1.0"/>
<meta name="csrf-token" content="{{ session.csrf_token }}">
<script>
    (function() {
        const originalFetch = window.fetch;
        window.fetch = async function(url, options = {}) {
            if (options.method && ['POST', 'PUT', 'DELETE', 'PATCH'].includes(options.method.toUpperCase())) {
                options.headers = options.headers || {};
                const token = document.querySelector('meta[name="csrf-token"]')?.getAttribute('content');
                if (token) {
                    options.headers['X-CSRFToken'] = token;
                }
            }
            return originalFetch(url, options);
        };
    })();
</script>
<title>Leave Management — AttendanceMS</title>
<link href="https://fonts.googleapis.com/css2?family=DM+Serif+Display:ital@0;1&family=DM+Sans:wght@300;400;500;600&display=swap" rel="stylesheet"/>
<style>
:root{
  --navy:#0f2342;--navy2:#1a3a5c;--navy3:#243b55;
  --accent:#2563eb;--accent2:#3b82f6;--accent3:#60a5fa;
  --gold:#f59e0b;--gold2:#fbbf24;
  --green:#059669;--green2:#d1fae5;
  --red:#dc2626;--red2:#fee2e2;
  --amber:#d97706;--amber2:#fef3c7;
  --white:#ffffff;--grey:#f8fafc;--grey2:#e2e8f0;--grey3:#94a3b8;
  --text:#1e293b;--text2:#475569;
  --radius:12px;--radius2:8px;
  --shadow:0 4px 24px rgba(15,35,66,.12);
  --shadow2:0 2px 8px rgba(15,35,66,.08);
}
*{box-sizing:border-box;margin:0;padding:0;}
body{font-family:'DM Sans',sans-serif;background:var(--grey);color:var(--text);min-height:100vh;}

/* ── LOGIN ── */
.login-wrap{min-height:100vh;display:flex;align-items:center;justify-content:center;
  background:linear-gradient(135deg,var(--navy) 0%,var(--navy2) 55%,#1e4080 100%);}
.login-card{background:var(--white);border-radius:20px;padding:48px 44px;width:420px;
  box-shadow:0 32px 80px rgba(0,0,0,.3);}
.login-logo{display:flex;align-items:center;gap:12px;margin-bottom:32px;}
.login-logo .icon{width:48px;height:48px;background:var(--accent);border-radius:12px;
  display:flex;align-items:center;justify-content:center;font-size:22px;}
.login-logo h1{font-family:'DM Serif Display',serif;font-size:22px;color:var(--navy);}
.login-logo p{font-size:12px;color:var(--grey3);margin-top:2px;}
.login-tabs{display:flex;gap:0;background:var(--grey);border-radius:10px;padding:4px;margin-bottom:28px;}
.login-tab{flex:1;padding:10px;border:none;background:none;border-radius:8px;cursor:pointer;
  font-family:'DM Sans',sans-serif;font-size:14px;font-weight:500;color:var(--text2);transition:.2s;}
.login-tab.active{background:var(--white);color:var(--navy);box-shadow:var(--shadow2);}
.field{margin-bottom:18px;}
.field label{display:block;font-size:12px;font-weight:600;color:var(--text2);
  text-transform:uppercase;letter-spacing:.5px;margin-bottom:6px;}
.field input,.field select,.field textarea{width:100%;padding:12px 14px;
  border:1.5px solid var(--grey2);border-radius:var(--radius2);
  font-family:'DM Sans',sans-serif;font-size:14px;color:var(--text);
  background:var(--grey);transition:border .2s,box-shadow .2s;outline:none;}
.field input:focus,.field select:focus,.field textarea:focus{
  border-color:var(--accent);background:var(--white);
  box-shadow:0 0 0 3px rgba(37,99,235,.12);}
.field textarea{resize:vertical;min-height:90px;}
.btn{width:100%;padding:14px;border:none;border-radius:var(--radius2);cursor:pointer;
  font-family:'DM Sans',sans-serif;font-size:15px;font-weight:600;
  transition:.2s;letter-spacing:.3px;}
.btn-primary{background:var(--accent);color:var(--white);}
.btn-primary:hover{background:#1d4ed8;transform:translateY(-1px);box-shadow:0 6px 20px rgba(37,99,235,.35);}
.hint{font-size:12px;color:var(--grey3);text-align:center;margin-top:16px;}
.hint span{color:var(--accent);font-weight:500;}

/* ── APP SHELL ── */
.app{display:flex;min-height:100vh;}
.sidebar{width:260px;background:var(--navy);display:flex;flex-direction:column;
  flex-shrink:0;position:fixed;top:0;left:0;bottom:0;z-index:100;}
.sidebar-top{padding:28px 24px 20px;}
.sidebar-logo{display:flex;align-items:center;gap:10px;margin-bottom:28px;}
.sidebar-logo .ic{width:38px;height:38px;background:var(--accent);border-radius:10px;
  display:flex;align-items:center;justify-content:center;font-size:18px;}
.sidebar-logo h2{font-family:'DM Serif Display',serif;color:var(--white);font-size:17px;line-height:1.2;}
.sidebar-logo small{color:var(--accent3);font-size:11px;}
.sidebar-user{background:rgba(255,255,255,.07);border-radius:10px;padding:12px 14px;margin-bottom:20px;}
.sidebar-user .role-badge{display:inline-block;padding:2px 10px;border-radius:20px;
  font-size:11px;font-weight:600;text-transform:uppercase;letter-spacing:.5px;margin-bottom:6px;}
.badge-admin{background:var(--gold);color:#78350f;}
.badge-student{background:var(--accent2);color:var(--white);}
.sidebar-user .uname{font-size:14px;font-weight:600;color:var(--white);}
.sidebar-user .uid{font-size:12px;color:var(--accent3);}
nav a{display:flex;align-items:center;gap:12px;padding:11px 24px;
  color:rgba(255,255,255,.65);text-decoration:none;font-size:14px;font-weight:500;
  transition:.15s;border-left:3px solid transparent;}
nav a:hover{color:var(--white);background:rgba(255,255,255,.06);}
nav a.active{color:var(--white);background:rgba(37,99,235,.25);border-left-color:var(--accent2);}
nav a .nav-ic{font-size:16px;width:20px;text-align:center;}
.sidebar-footer{margin-top:auto;padding:20px 24px;}
.btn-logout{width:100%;padding:10px;background:rgba(220,38,38,.15);border:1px solid rgba(220,38,38,.3);
  border-radius:8px;color:#fca5a5;font-family:'DM Sans',sans-serif;
  font-size:13px;font-weight:600;cursor:pointer;transition:.2s;}
.btn-logout:hover{background:rgba(220,38,38,.3);}

.main{margin-left:260px;flex:1;padding:32px 36px;}
.page{display:none;}
.page.active{display:block;}

/* ── TOPBAR ── */
.topbar{display:flex;align-items:center;justify-content:space-between;margin-bottom:32px;}
.topbar h2{font-family:'DM Serif Display',serif;font-size:28px;color:var(--navy);}
.topbar p{font-size:14px;color:var(--text2);margin-top:2px;}

/* ── STAT CARDS ── */
.stats{display:grid;grid-template-columns:repeat(4,1fr);gap:16px;margin-bottom:28px;}
.stat-card{background:var(--white);border-radius:var(--radius);padding:22px;
  box-shadow:var(--shadow2);border-top:4px solid transparent;transition:.2s;}
.stat-card:hover{transform:translateY(-2px);box-shadow:var(--shadow);}
.stat-card.c-total{border-color:var(--accent);}
.stat-card.c-pending{border-color:var(--amber);}
.stat-card.c-approved{border-color:var(--green);}
.stat-card.c-rejected{border-color:var(--red);}
.stat-num{font-family:'DM Serif Display',serif;font-size:36px;color:var(--navy);}
.stat-label{font-size:13px;color:var(--text2);font-weight:500;margin-top:4px;}

/* ── TABLE ── */
.card{background:var(--white);border-radius:var(--radius);box-shadow:var(--shadow2);overflow:hidden;margin-bottom:24px;}
.card-head{padding:18px 24px;display:flex;align-items:center;justify-content:space-between;
  border-bottom:1px solid var(--grey2);}
.card-head h3{font-size:16px;font-weight:600;color:var(--navy);}
.filters{display:flex;gap:10px;align-items:center;}
.filters select,.filters input{padding:8px 12px;border:1.5px solid var(--grey2);border-radius:8px;
  font-family:'DM Sans',sans-serif;font-size:13px;color:var(--text);background:var(--grey);outline:none;}
.filters select:focus,.filters input:focus{border-color:var(--accent);}
.tbl-wrap{overflow-x:auto;}
table{width:100%;border-collapse:collapse;}
th{padding:11px 16px;background:var(--grey);font-size:12px;font-weight:600;
  color:var(--text2);text-transform:uppercase;letter-spacing:.4px;text-align:left;
  border-bottom:1px solid var(--grey2);}
td{padding:13px 16px;font-size:13.5px;border-bottom:1px solid #f1f5f9;color:var(--text);}
tr:last-child td{border-bottom:none;}
tr:hover td{background:#f8faff;}
.status-pill{display:inline-flex;align-items:center;gap:5px;padding:4px 12px;
  border-radius:20px;font-size:12px;font-weight:600;}
.status-pill::before{content:'';width:7px;height:7px;border-radius:50%;}
.pill-pending{background:var(--amber2);color:var(--amber);}
.pill-pending::before{background:var(--amber);}
.pill-approved{background:var(--green2);color:var(--green);}
.pill-approved::before{background:var(--green);}
.pill-rejected{background:var(--red2);color:var(--red);}
.pill-rejected::before{background:var(--red);}
.btn-sm{padding:6px 14px;border:none;border-radius:6px;cursor:pointer;
  font-family:'DM Sans',sans-serif;font-size:12px;font-weight:600;transition:.15s;}
.btn-approve{background:var(--green2);color:var(--green);}
.btn-approve:hover{background:var(--green);color:var(--white);}
.btn-reject{background:var(--red2);color:var(--red);}
.btn-reject:hover{background:var(--red);color:var(--white);}
.btn-view{background:#e0f2fe;color:#0369a1;}
.btn-view:hover{background:#0369a1;color:var(--white);}
.empty-state{text-align:center;padding:60px 24px;color:var(--text2);}
.empty-state .ei{font-size:48px;margin-bottom:12px;opacity:.4;}
.empty-state p{font-size:14px;}

/* ── APPLY FORM ── */
.apply-wrap{max-width:680px;}
.apply-card{background:var(--white);border-radius:var(--radius);padding:32px;
  box-shadow:var(--shadow2);}
.apply-card h3{font-family:'DM Serif Display',serif;font-size:22px;color:var(--navy);margin-bottom:6px;}
.apply-card .sub{font-size:13px;color:var(--text2);margin-bottom:28px;}
.grid2{display:grid;grid-template-columns:1fr 1fr;gap:18px;}
.days-display{background:var(--grey);border-radius:8px;padding:12px 16px;
  font-size:14px;color:var(--text2);margin-top:6px;}
.days-display span{font-weight:700;color:var(--accent);font-size:18px;}
.btn-submit{background:linear-gradient(135deg,var(--accent),var(--navy2));color:var(--white);
  padding:14px 32px;border:none;border-radius:var(--radius2);cursor:pointer;
  font-family:'DM Sans',sans-serif;font-size:15px;font-weight:600;
  transition:.2s;letter-spacing:.3px;margin-top:8px;}
.btn-submit:hover{transform:translateY(-2px);box-shadow:0 8px 24px rgba(37,99,235,.4);}

/* ── MODAL ── */
.overlay{position:fixed;inset:0;background:rgba(15,35,66,.6);z-index:500;
  display:none;align-items:center;justify-content:center;backdrop-filter:blur(3px);}
.overlay.open{display:flex;}
.modal{background:var(--white);border-radius:16px;width:520px;max-width:95vw;
  box-shadow:0 32px 80px rgba(0,0,0,.3);overflow:hidden;}
.modal-head{padding:22px 28px;background:var(--navy);display:flex;align-items:center;justify-content:space-between;}
.modal-head h4{font-family:'DM Serif Display',serif;color:var(--white);font-size:18px;}
.modal-close{background:none;border:none;color:rgba(255,255,255,.6);font-size:22px;cursor:pointer;transition:.2s;}
.modal-close:hover{color:var(--white);}
.modal-body{padding:28px;}
.detail-grid{display:grid;grid-template-columns:1fr 1fr;gap:14px;margin-bottom:22px;}
.detail-item label{font-size:11px;font-weight:600;color:var(--text2);text-transform:uppercase;letter-spacing:.4px;}
.detail-item p{font-size:14px;color:var(--navy);margin-top:3px;font-weight:500;}
.detail-item.full{grid-column:1/-1;}
.review-actions{display:flex;gap:12px;margin-top:4px;}
.review-actions .btn{flex:1;padding:12px;}

/* ── TOAST ── */
.toast-wrap{position:fixed;top:20px;right:24px;z-index:999;display:flex;flex-direction:column;gap:10px;}
.toast{background:var(--white);border-radius:10px;padding:14px 18px;box-shadow:var(--shadow);
  display:flex;align-items:center;gap:10px;font-size:14px;font-weight:500;
  animation:slideIn .3s ease;border-left:4px solid var(--accent);min-width:280px;}
.toast.success{border-left-color:var(--green);}
.toast.error{border-left-color:var(--red);}
.toast.warning{border-left-color:var(--amber);}
@keyframes slideIn{from{opacity:0;transform:translateX(40px);}to{opacity:1;transform:translateX(0);}}

/* ── RESPONSIVE ── */
@media(max-width:900px){
  .sidebar{width:220px;}
  .main{margin-left:220px;padding:24px;}
  .stats{grid-template-columns:repeat(2,1fr);}
}
</style>
</head>
<body>

<!-- ═══════════════ LOGIN SCREEN ═══════════════ -->
<div class="login-wrap" id="loginScreen">
  <div class="login-card">
    <div class="login-logo">
      <div class="icon">📋</div>
      <div>
        <h1>LeaveMS</h1>
        <p>Leave Management System</p>
      </div>
    </div>
    <div class="login-tabs">
      <button class="login-tab active" onclick="switchLoginTab('student')">Student</button>
      <button class="login-tab" onclick="switchLoginTab('admin')">Admin</button>
    </div>
    <div id="loginForm">
      <div class="field">
        <label id="loginLabel">Student ID</label>
        <input type="text" id="loginUsername" placeholder="Enter your Student ID"/>
      </div>
      <div class="field">
        <label>Password</label>
        <input type="password" id="loginPassword" placeholder="••••••••"/>
      </div>
      <button class="btn btn-primary" onclick="doLogin()">Sign In →</button>
      <p class="hint" id="loginHint">
        Default password is your <span>Student ID</span>
      </p>
    </div>
  </div>
</div>

<!-- ═══════════════ APP SHELL ═══════════════ -->
<div class="app" id="appShell" style="display:none">
  <aside class="sidebar">
    <div class="sidebar-top">
      <div class="sidebar-logo">
        <div class="ic">📋</div>
        <div>
          <h2>LeaveMS</h2>
          <small>Management System</small>
        </div>
      </div>
      <div class="sidebar-user">
        <div class="role-badge" id="roleBadge">—</div>
        <div class="uname" id="sidebarName">—</div>
        <div class="uid" id="sidebarId"></div>
      </div>
    </div>

    <!-- STUDENT NAV -->
    <nav id="studentNav" style="display:none">
      <a href="#" class="active" onclick="showPage('apply')" data-page="apply">
        <span class="nav-ic">✏️</span> Apply for Leave
      </a>
      <a href="#" onclick="showPage('myLeaves')" data-page="myLeaves">
        <span class="nav-ic">📄</span> My Applications
      </a>
    </nav>

    <!-- ADMIN NAV -->
    <nav id="adminNav" style="display:none">
      <a href="#" class="active" onclick="showPage('adminDash')" data-page="adminDash">
        <span class="nav-ic">📊</span> Dashboard
      </a>
      <a href="#" onclick="showPage('allLeaves')" data-page="allLeaves">
        <span class="nav-ic">📋</span> All Applications
      </a>
      <a href="#" onclick="showPage('studentSum')" data-page="studentSum">
        <span class="nav-ic">👥</span> Student Summary
      </a>
    </nav>

    <div class="sidebar-footer">
      <button class="btn-logout" onclick="doLogout()">⬅ Sign Out</button>
    </div>
  </aside>

  <main class="main">
    <!-- ─── APPLY LEAVE (Student) ─── -->
    <div class="page active" id="page-apply">
      <div class="topbar">
        <div>
          <h2>Apply for Leave</h2>
          <p>Submit a new leave application</p>
        </div>
      </div>
      <div class="apply-wrap">
        <div class="apply-card">
          <h3>Leave Application Form</h3>
          <p class="sub">Fill in the details below. Your request will be reviewed by the admin.</p>
          <div class="field">
            <label>Leave Type</label>
            <select id="leaveType">
              <option value="">Select leave type…</option>
              <option value="Medical Leave">🏥 Medical Leave</option>
              <option value="Family Emergency">🏠 Family Emergency</option>
              <option value="Personal Leave">🙋 Personal Leave</option>
              <option value="Educational Event">🎓 Educational Event</option>
              <option value="Bereavement">🕊️ Bereavement</option>
              <option value="Other">📝 Other</option>
            </select>
          </div>
          <div class="grid2">
            <div class="field">
              <label>From Date</label>
              <input type="date" id="fromDate" onchange="calcDays()"/>
            </div>
            <div class="field">
              <label>To Date</label>
              <input type="date" id="toDate" onchange="calcDays()"/>
            </div>
          </div>
          <div class="days-display" id="daysDisplay" style="display:none">
            Duration: <span id="daysCount">0</span> day(s)
          </div>
          <div class="field" style="margin-top:18px">
            <label>Reason for Leave</label>
            <textarea id="leaveReason" placeholder="Please describe the reason for your leave request…"></textarea>
          </div>
          <button class="btn-submit" onclick="applyLeave()">Submit Application →</button>
        </div>
      </div>
    </div>

    <!-- ─── MY LEAVES (Student) ─── -->
    <div class="page" id="page-myLeaves">
      <div class="topbar">
        <div>
          <h2>My Applications</h2>
          <p>Track status of your leave requests</p>
        </div>
      </div>
      <div class="card">
        <div class="card-head">
          <h3>Leave History</h3>
          <div class="filters">
            <select id="myStatusFilter" onchange="renderMyLeaves()">
              <option value="">All Status</option>
              <option value="Pending">Pending</option>
              <option value="Approved">Approved</option>
              <option value="Rejected">Rejected</option>
            </select>
          </div>
        </div>
        <div class="tbl-wrap">
          <table>
            <thead>
              <tr>
                <th>Leave ID</th><th>Type</th><th>From</th><th>To</th>
                <th>Days</th><th>Status</th><th>Remarks</th><th>Applied On</th>
              </tr>
            </thead>
            <tbody id="myLeavesTbody"></tbody>
          </table>
        </div>
      </div>
    </div>

    <!-- ─── ADMIN DASHBOARD ─── -->
    <div class="page" id="page-adminDash">
      <div class="topbar">
        <div>
          <h2>Dashboard</h2>
          <p>Overview of all leave applications</p>
        </div>
        <button class="btn-submit" style="padding:10px 22px;font-size:14px" onclick="showPage('allLeaves')">
          View All →
        </button>
      </div>
      <div class="stats">
        <div class="stat-card c-total">
          <div class="stat-num" id="st-total">0</div>
          <div class="stat-label">Total Applications</div>
        </div>
        <div class="stat-card c-pending">
          <div class="stat-num" id="st-pending">0</div>
          <div class="stat-label">Pending Review</div>
        </div>
        <div class="stat-card c-approved">
          <div class="stat-num" id="st-approved">0</div>
          <div class="stat-label">Approved</div>
        </div>
        <div class="stat-card c-rejected">
          <div class="stat-num" id="st-rejected">0</div>
          <div class="stat-label">Rejected</div>
        </div>
      </div>
      <!-- Recent pending -->
      <div class="card">
        <div class="card-head"><h3>⏳ Pending Applications</h3></div>
        <div class="tbl-wrap">
          <table>
            <thead>
              <tr><th>Leave ID</th><th>Student</th><th>Class</th><th>Type</th>
                <th>From</th><th>To</th><th>Days</th><th>Action</th></tr>
            </thead>
            <tbody id="pendingTbody"></tbody>
          </table>
        </div>
      </div>
    </div>

    <!-- ─── ALL LEAVES (Admin) ─── -->
    <div class="page" id="page-allLeaves">
      <div class="topbar">
        <div>
          <h2>All Applications</h2>
          <p>Review and manage leave requests</p>
        </div>
      </div>
      <div class="card">
        <div class="card-head">
          <h3>Applications</h3>
          <div class="filters">
            <input type="text" id="searchFilter" placeholder="Search student…" oninput="renderAllLeaves()"/>
            <select id="statusFilter" onchange="renderAllLeaves()">
              <option value="">All Status</option>
              <option value="Pending">Pending</option>
              <option value="Approved">Approved</option>
              <option value="Rejected">Rejected</option>
            </select>
          </div>
        </div>
        <div class="tbl-wrap">
          <table>
            <thead>
              <tr><th>Leave ID</th><th>Student</th><th>Class</th><th>Type</th>
                <th>From</th><th>To</th><th>Days</th><th>Status</th><th>Action</th></tr>
            </thead>
            <tbody id="allLeavesTbody"></tbody>
          </table>
        </div>
      </div>
    </div>

    <!-- ─── STUDENT SUMMARY (Admin) ─── -->
    <div class="page" id="page-studentSum">
      <div class="topbar">
        <div><h2>Student Summary</h2><p>Leave statistics per student</p></div>
      </div>
      <div class="card">
        <div class="card-head"><h3>Per-Student Leave Stats</h3></div>
        <div class="tbl-wrap">
          <table>
            <thead>
              <tr><th>Student ID</th><th>Name</th><th>Total</th>
                <th>Pending</th><th>Approved</th><th>Rejected</th><th>Approved Days</th></tr>
            </thead>
            <tbody id="studentSumTbody"></tbody>
          </table>
        </div>
      </div>
    </div>
  </main>
</div>

<!-- ─── REVIEW MODAL ─── -->
<div class="overlay" id="reviewModal">
  <div class="modal">
    <div class="modal-head">
      <h4>Review Application</h4>
      <button class="modal-close" onclick="closeModal()">✕</button>
    </div>
    <div class="modal-body">
      <div class="detail-grid">
        <div class="detail-item"><label>Leave ID</label><p id="m-lid"></p></div>
        <div class="detail-item"><label>Student</label><p id="m-sname"></p></div>
        <div class="detail-item"><label>Class</label><p id="m-class"></p></div>
        <div class="detail-item"><label>Leave Type</label><p id="m-type"></p></div>
        <div class="detail-item"><label>From</label><p id="m-from"></p></div>
        <div class="detail-item"><label>To</label><p id="m-to"></p></div>
        <div class="detail-item"><label>Days</label><p id="m-days"></p></div>
        <div class="detail-item"><label>Status</label><p id="m-status"></p></div>
        <div class="detail-item full"><label>Reason</label><p id="m-reason"></p></div>
      </div>
      <div class="field" id="reviewFields">
        <label>Admin Remarks (optional)</label>
        <textarea id="adminRemarks" placeholder="Add remarks or notes for the student…" style="min-height:70px"></textarea>
      </div>
      <div class="review-actions" id="reviewBtns">
        <button class="btn btn-approve btn-sm" style="padding:11px;font-size:14px" onclick="submitReview('Approved')">✓ Approve</button>
        <button class="btn btn-reject btn-sm" style="padding:11px;font-size:14px" onclick="submitReview('Rejected')">✗ Reject</button>
      </div>
    </div>
  </div>
</div>

<!-- TOAST CONTAINER -->
<div class="toast-wrap" id="toastWrap"></div>

<script>
// ═══ STATE ═══
let currentRole = null, currentLeaveId = null;
let allLeavesData = [], myLeavesData = [];

// ═══ TOAST ═══
function toast(msg, type='success'){
  const wrap = document.getElementById('toastWrap');
  const el = document.createElement('div');
  el.className = `toast ${type}`;
  const ic = type==='success'?'✓':type==='error'?'✗':'⚠';
  el.innerHTML = `<span style="font-size:16px">${ic}</span> ${msg}`;
  wrap.appendChild(el);
  setTimeout(()=>el.remove(), 3800);
}

// ═══ LOGIN ═══
let loginTabRole = 'student';
function switchLoginTab(role){
  loginTabRole = role;
  document.querySelectorAll('.login-tab').forEach((t,i)=>
    t.classList.toggle('active', (i===0&&role==='student')||(i===1&&role==='admin')));
  document.getElementById('loginLabel').textContent = role==='admin'?'Username':'Student ID';
  document.getElementById('loginUsername').placeholder = role==='admin'?'admin':'Enter your Student ID';
  document.getElementById('loginHint').innerHTML = role==='admin'
    ? 'Contact system administrator for access'
    : 'Default password is your <span>Student ID</span>';
}

async function doLogin(){
  const u = document.getElementById('loginUsername').value.trim();
  const p = document.getElementById('loginPassword').value.trim();
  if(!u||!p){toast('Please fill in all fields','error');return;}
  const res = await fetch('/leave/api/auth/login',{
    method:'POST',headers:{'Content-Type':'application/json'},
    body:JSON.stringify({username:u,password:p})});
  const d = await res.json();
  if(d.success){
    currentRole = d.role;
    initShell(d);
    document.getElementById('loginScreen').style.display='none';
    document.getElementById('appShell').style.display='flex';
    toast(`Welcome, ${d.name}!`);
  } else {
    toast(d.message||'Login failed','error');
  }
}

document.addEventListener('keydown',e=>{if(e.key==='Enter')doLogin();});

async function doLogout(){
  await fetch('/leave/api/auth/logout',{method:'POST'});
  document.getElementById('appShell').style.display='none';
  document.getElementById('loginScreen').style.display='flex';
  document.getElementById('loginUsername').value='';
  document.getElementById('loginPassword').value='';
  currentRole=null; toast('Signed out successfully');
}

function initShell(d){
  const badge = document.getElementById('roleBadge');
  badge.textContent = d.role==='admin'?'Administrator':'Student';
  badge.className = 'role-badge '+(d.role==='admin'?'badge-admin':'badge-student');
  document.getElementById('sidebarName').textContent = d.name;
  document.getElementById('sidebarId').textContent = d.role==='student'?`ID: ${d.student_id}`:'';
  document.getElementById('studentNav').style.display = d.role==='student'?'block':'none';
  document.getElementById('adminNav').style.display   = d.role==='admin'?'block':'none';
  if(d.role==='admin'){
    showPage('adminDash'); loadAdminDash();
  } else {
    showPage('apply');
  }
}

// ═══ NAVIGATION ═══
function showPage(name){
  document.querySelectorAll('.page').forEach(p=>p.classList.remove('active'));
  document.getElementById('page-'+name).classList.add('active');
  document.querySelectorAll('nav a').forEach(a=>{
    a.classList.toggle('active', a.dataset.page===name);
  });
  if(name==='myLeaves')    loadMyLeaves();
  if(name==='allLeaves')   loadAllLeaves();
  if(name==='adminDash')   loadAdminDash();
  if(name==='studentSum')  loadStudentSummary();
  return false;
}

// ═══ DAYS CALC ═══
function calcDays(){
  const f=document.getElementById('fromDate').value;
  const t=document.getElementById('toDate').value;
  if(f&&t){
    const d = Math.max(1, Math.round((new Date(t)-new Date(f))/(86400000))+1);
    document.getElementById('daysCount').textContent = d;
    document.getElementById('daysDisplay').style.display='block';
  }
}

// ═══ APPLY LEAVE ═══
async function applyLeave(){
  const body = {
    leave_type: document.getElementById('leaveType').value,
    from_date:  document.getElementById('fromDate').value,
    to_date:    document.getElementById('toDate').value,
    reason:     document.getElementById('leaveReason').value.trim()
  };
  if(!body.leave_type||!body.from_date||!body.to_date||!body.reason){
    toast('Please fill in all fields','error'); return;
  }
  if(body.from_date > body.to_date){
    toast('"From" date cannot be after "To" date','error'); return;
  }
  const res = await fetch('/leave/api/apply',{
    method:'POST',headers:{'Content-Type':'application/json'},body:JSON.stringify(body)});
  const d = await res.json();
  if(d.success){
    toast(d.message);
    document.getElementById('leaveType').value='';
    document.getElementById('fromDate').value='';
    document.getElementById('toDate').value='';
    document.getElementById('leaveReason').value='';
    document.getElementById('daysDisplay').style.display='none';
  } else {
    toast(d.message,'error');
  }
}

// ═══ MY LEAVES ═══
async function loadMyLeaves(){
  const res = await fetch('/leave/api/my_leaves');
  const d = await res.json();
  if(d.success) myLeavesData = d.leaves;
  renderMyLeaves();
}

function renderMyLeaves(){
  const filter = document.getElementById('myStatusFilter').value;
  const data = myLeavesData.filter(r=>!filter||r.status===filter);
  const tbody = document.getElementById('myLeavesTbody');
  if(!data.length){
    tbody.innerHTML=`<tr><td colspan="8">
      <div class="empty-state"><div class="ei">📭</div>
      <p>No leave applications found.</p></div></td></tr>`;
    return;
  }
  tbody.innerHTML = data.map(r=>`
    <tr>
      <td><code style="background:#f1f5f9;padding:3px 7px;border-radius:4px;font-size:12px">${r.leave_id}</code></td>
      <td>${r.leave_type}</td>
      <td>${r.from_date}</td>
      <td>${r.to_date}</td>
      <td><b>${r.days}</b></td>
      <td>${pillHTML(r.status)}</td>
      <td style="max-width:180px;font-size:12px;color:var(--text2)">${r.remarks||'—'}</td>
      <td style="font-size:12px;color:var(--grey3)">${r.applied_on}</td>
    </tr>`).join('');
}

// ═══ ADMIN DASHBOARD ═══
async function loadAdminDash(){
  const res = await fetch('/leave/api/all_leaves');
  const d = await res.json();
  if(!d.success) return;
  allLeavesData = d.leaves;
  const s = d.stats;
  document.getElementById('st-total').textContent   = s.total;
  document.getElementById('st-pending').textContent = s.pending;
  document.getElementById('st-approved').textContent= s.approved;
  document.getElementById('st-rejected').textContent= s.rejected;
  const pending = d.leaves.filter(r=>r.status==='Pending');
  const tbody = document.getElementById('pendingTbody');
  if(!pending.length){
    tbody.innerHTML=`<tr><td colspan="8"><div class="empty-state">
      <div class="ei">🎉</div><p>No pending applications!</p></div></td></tr>`;
    return;
  }
  tbody.innerHTML = pending.map(r=>`
    <tr>
      <td><code style="background:#f1f5f9;padding:3px 7px;border-radius:4px;font-size:12px">${r.leave_id}</code></td>
      <td><b>${r.student_name}</b><br><span style="font-size:11px;color:var(--grey3)">${r.student_id}</span></td>
      <td>${r.class||'—'}</td>
      <td>${r.leave_type}</td>
      <td>${r.from_date}</td><td>${r.to_date}</td>
      <td><b>${r.days}</b></td>
      <td><button class="btn-sm btn-view" onclick="openReview('${r.leave_id}')">Review</button></td>
    </tr>`).join('');
}

// ═══ ALL LEAVES ═══
async function loadAllLeaves(){
  const res = await fetch('/leave/api/all_leaves');
  const d = await res.json();
  if(d.success) allLeavesData = d.leaves;
  renderAllLeaves();
}

function renderAllLeaves(){
  const search = (document.getElementById('searchFilter').value||'').toLowerCase();
  const filter = document.getElementById('statusFilter').value;
  const data = allLeavesData.filter(r=>{
    const matchSearch = !search ||
      r.student_name.toLowerCase().includes(search) ||
      r.student_id.toLowerCase().includes(search);
    const matchStatus = !filter || r.status===filter;
    return matchSearch && matchStatus;
  });
  const tbody = document.getElementById('allLeavesTbody');
  if(!data.length){
    tbody.innerHTML=`<tr><td colspan="9"><div class="empty-state">
      <div class="ei">📭</div><p>No applications found.</p></div></td></tr>`;
    return;
  }
  tbody.innerHTML = data.map(r=>`
    <tr>
      <td><code style="background:#f1f5f9;padding:3px 7px;border-radius:4px;font-size:12px">${r.leave_id}</code></td>
      <td><b>${r.student_name}</b><br><span style="font-size:11px;color:var(--grey3)">${r.student_id}</span></td>
      <td>${r.class||'—'}</td>
      <td>${r.leave_type}</td>
      <td>${r.from_date}</td><td>${r.to_date}</td>
      <td><b>${r.days}</b></td>
      <td>${pillHTML(r.status)}</td>
      <td>
        ${r.status==='Pending'
          ? `<button class="btn-sm btn-view" onclick="openReview('${r.leave_id}')">Review</button>`
          : `<button class="btn-sm" style="background:var(--grey);color:var(--text2)" onclick="openView('${r.leave_id}')">View</button>`}
      </td>
    </tr>`).join('');
}

// ═══ STUDENT SUMMARY ═══
async function loadStudentSummary(){
  const res = await fetch('/leave/api/student_summary');
  const d = await res.json();
  if(!d.success) return;
  const tbody = document.getElementById('studentSumTbody');
  if(!d.summary.length){
    tbody.innerHTML=`<tr><td colspan="7"><div class="empty-state">
      <div class="ei">📭</div><p>No data yet.</p></div></td></tr>`;
    return;
  }
  tbody.innerHTML = d.summary.map(s=>`
    <tr>
      <td><code style="background:#f1f5f9;padding:3px 7px;border-radius:4px;font-size:12px">${s.student_id}</code></td>
      <td><b>${s.name}</b></td>
      <td>${s.total}</td>
      <td>${s.pending>0?`<span style="color:var(--amber);font-weight:700">${s.pending}</span>`:s.pending}</td>
      <td>${s.approved>0?`<span style="color:var(--green);font-weight:700">${s.approved}</span>`:s.approved}</td>
      <td>${s.rejected>0?`<span style="color:var(--red);font-weight:700">${s.rejected}</span>`:s.rejected}</td>
      <td><b>${s.total_days}</b> days</td>
    </tr>`).join('');
}

// ═══ REVIEW MODAL ═══
function openReview(lid){
  const r = allLeavesData.find(x=>x.leave_id===lid);
  if(!r) return;
  currentLeaveId = lid;
  document.getElementById('m-lid').textContent   = r.leave_id;
  document.getElementById('m-sname').textContent = `${r.student_name} (${r.student_id})`;
  document.getElementById('m-class').textContent = r.class||'—';
  document.getElementById('m-type').textContent  = r.leave_type;
  document.getElementById('m-from').textContent  = r.from_date;
  document.getElementById('m-to').textContent    = r.to_date;
  document.getElementById('m-days').textContent  = r.days+' day(s)';
  document.getElementById('m-status').innerHTML  = pillHTML(r.status);
  document.getElementById('m-reason').textContent= r.reason;
  document.getElementById('adminRemarks').value  = '';
  document.getElementById('reviewFields').style.display = r.status==='Pending'?'':'none';
  document.getElementById('reviewBtns').style.display   = r.status==='Pending'?'flex':'none';
  document.getElementById('reviewModal').classList.add('open');
}

function openView(lid){ openReview(lid); }
function closeModal(){ document.getElementById('reviewModal').classList.remove('open'); }

async function submitReview(action){
  const remarks = document.getElementById('adminRemarks').value.trim();
  const res = await fetch('/leave/api/review',{
    method:'POST',headers:{'Content-Type':'application/json'},
    body:JSON.stringify({leave_id:currentLeaveId, action, remarks})});
  const d = await res.json();
  if(d.success){
    toast(d.message, action==='Approved'?'success':'warning');
    closeModal();
    loadAdminDash();
    // refresh all leaves if on that page
    const allPage = document.getElementById('page-allLeaves');
    if(allPage.classList.contains('active')) loadAllLeaves();
  } else {
    toast(d.message,'error');
  }
}

// ═══ HELPERS ═══
function pillHTML(status){
  const cls = status==='Approved'?'pill-approved':status==='Rejected'?'pill-rejected':'pill-pending';
  return `<span class="status-pill ${cls}">${status}</span>`;
}

// ═══ INIT: check session ═══
(async()=>{
  const r = await fetch('/leave/api/auth/me');
  const d = await r.json();
  if(d.logged_in){
    currentRole = d.role;
    document.getElementById('loginScreen').style.display='none';
    document.getElementById('appShell').style.display='flex';
    initShell(d);
  }
})();
</script>
</body>
</html>"""

if __name__ == "__main__":
    get_leave_wb()
    print("\n+------------------------------------------+")
    print("|   Leave Management System                |")
    print("|   http://localhost:5001/leave            |")
    print("+------------------------------------------+")
    print("|  Admin  ->  admin / admin123              |")
    print("|  Student -> <student_id> / <student_id>   |")
    print("+------------------------------------------+\n")
    app.run(debug=False, port=5001)
