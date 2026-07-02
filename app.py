from flask import Flask, render_template, request, jsonify, send_file, Response, session, redirect, url_for, flash
from datetime import datetime, timedelta
import sqlite3
import openpyxl
import os
import io
import secrets
import re
import html
import time
from collections import defaultdict
from functools import wraps
from werkzeug.security import generate_password_hash, check_password_hash

app = Flask(__name__)
app.secret_key = secrets.token_hex(24)

# Secure Session Cookies
app.config.update(
    SESSION_COOKIE_HTTPONLY=True,
    SESSION_COOKIE_SECURE=True,
    SESSION_COOKIE_SAMESITE='Lax'
)

DB_FILE = "attendance.db"
EXCEL_FILE = "attendance_data.xlsx"
LEAVE_FILE = "leave_data.xlsx"

LOGIN_HOUR = 7    # 7 AM
LATE_MIN = 30     # 7:30 AM
LOGOUT_HOUR = 12  # 12 PM

# ─────────────────────────────────────────────────────────────
#  SECURITY HELPERS, RATE LIMITING & CSRF PROTECTION
# ─────────────────────────────────────────────────────────────

def is_valid_email(email):
    regex = r'^[a-zA-Z0-9._%+-]+@[a-zA-Z0-9.-]+\.[a-zA-Z]{2,}$'
    return bool(re.match(regex, email))

def is_valid_date(date_str):
    try:
        datetime.strptime(date_str, "%Y-%m-%d")
        return True
    except ValueError:
        return False

def is_alphanumeric(val):
    return bool(re.match(r'^[a-zA-Z0-9_-]+$', val))

class RateLimiter:
    def __init__(self, limit=5, window=60):
        self.limit = limit
        self.window = window
        self.attempts = defaultdict(list)
        
    def is_blocked(self, key):
        now = time.time()
        self.attempts[key] = [t for t in self.attempts[key] if now - t < self.window]
        if len(self.attempts[key]) >= self.limit:
            return True
        self.attempts[key].append(now)
        return False

login_limiter = RateLimiter(limit=5, window=60)

@app.before_request
def security_before_request():
    # 1. Initialize CSRF token if not present
    if 'csrf_token' not in session:
        session['csrf_token'] = secrets.token_hex(32)
        
    # 2. CSRF Protection for state-changing requests
    if request.method in ["POST", "PUT", "DELETE", "PATCH"]:
        csrf_token = request.headers.get("X-CSRFToken") or request.form.get("csrf_token")
        if not csrf_token or csrf_token != session.get("csrf_token"):
            if request.path.startswith('/api/'):
                return jsonify(success=False, error="CSRF", message="CSRF token missing or invalid."), 400
            return render_template('error.html', code=400, message="CSRF verification failed. The form submission was rejected."), 400

@app.after_request
def add_security_headers(response):
    response.headers['X-Frame-Options'] = 'SAMEORIGIN'
    response.headers['X-Content-Type-Options'] = 'nosniff'
    response.headers['X-XSS-Protection'] = '1; mode=block'
    response.headers['Referrer-Policy'] = 'strict-origin-when-cross-origin'
    
    # CSP: Allow self, Google Fonts, and chart.js via jsdelivr CDN
    csp = (
        "default-src 'self'; "
        "script-src 'self' 'unsafe-inline' https://cdn.jsdelivr.net; "
        "style-src 'self' 'unsafe-inline' https://fonts.googleapis.com; "
        "font-src 'self' https://fonts.gstatic.com; "
        "img-src 'self' data:; "
        "connect-src 'self';"
    )
    response.headers['Content-Security-Policy'] = csp
    return response

# ─────────────────────────────────────────────────────────────
#  DATABASE INIT & AUTO-MIGRATION
# ─────────────────────────────────────────────────────────────
def get_db_connection():
    conn = sqlite3.connect(DB_FILE)
    conn.row_factory = sqlite3.Row
    return conn

def sync_attendance_to_excel():
    """Rebuild the Attendance sheet in attendance_data.xlsx from SQLite so the
    Excel file always reflects the current database state."""
    try:
        # Load (or recreate) the workbook
        if os.path.exists(EXCEL_FILE):
            wb = openpyxl.load_workbook(EXCEL_FILE)
        else:
            wb = openpyxl.Workbook()

        # ── Attendance sheet ──────────────────────────────────────
        if "Attendance" in wb.sheetnames:
            # Remove and re-create so we start clean
            del wb["Attendance"]
        ws_att = wb.create_sheet("Attendance", 0)

        headers = ["Date", "Intern ID", "Intern Name",
                   "Login Time", "Logout Time", "Working Hours",
                   "Status", "Remarks"]
        ws_att.append(headers)

        # Style header row
        for col in range(1, len(headers) + 1):
            cell = ws_att.cell(row=1, column=col)
            cell.font = openpyxl.styles.Font(bold=True, color="FFFFFF")
            cell.fill = openpyxl.styles.PatternFill(
                start_color="1E3A5F", end_color="1E3A5F", fill_type="solid")
            cell.alignment = openpyxl.styles.Alignment(horizontal="center")

        # Fetch all attendance records joined with user name
        conn = get_db_connection()
        cursor = conn.cursor()
        cursor.execute("""
            SELECT a.date, a.user_id, u.name,
                   a.check_in, a.check_out,
                   a.working_hours, a.status, a.location
            FROM attendance a
            JOIN users u ON a.user_id = u.id
            ORDER BY a.date DESC, a.user_id
        """)
        rows = cursor.fetchall()
        conn.close()

        summary = {}  # date -> {total, present, absent, late, on_leave}
        for r in rows:
            check_in  = r["check_in"]  if r["check_in"]  and r["check_in"]  != "—" else "—"
            check_out = r["check_out"] if r["check_out"] and r["check_out"] != "—" else "—"
            hrs_str   = f"{r['working_hours']:.2f} hrs" if r["working_hours"] else "0.00 hrs"
            remarks   = r["location"] or ""
            ws_att.append([
                r["date"], r["user_id"], r["name"],
                check_in, check_out, hrs_str,
                r["status"], remarks
            ])

            # Accumulate summary
            d = r["date"]
            if d not in summary:
                summary[d] = {"total": 0, "present": 0, "absent": 0, "late": 0, "on_leave": 0}
            summary[d]["total"] += 1
            st = (r["status"] or "").lower()
            if st in ("present", "logged in"):
                summary[d]["present"] += 1
            elif st == "late":
                summary[d]["late"] += 1
                summary[d]["present"] += 1  # late counts as present
            elif st == "on leave":
                summary[d]["on_leave"] += 1
            else:
                summary[d]["absent"] += 1

        # Auto-fit column widths for Attendance sheet
        col_widths = [12, 12, 22, 12, 12, 15, 12, 30]
        for i, w in enumerate(col_widths, 1):
            ws_att.column_dimensions[
                openpyxl.utils.get_column_letter(i)].width = w

        # ── Summary sheet ─────────────────────────────────────────
        if "Summary" in wb.sheetnames:
            del wb["Summary"]
        ws_sum = wb.create_sheet("Summary")

        sum_headers = ["Date", "Total", "Present", "Absent",
                       "Late", "On Leave", "Attendance %"]
        ws_sum.append(sum_headers)
        for col in range(1, len(sum_headers) + 1):
            cell = ws_sum.cell(row=1, column=col)
            cell.font = openpyxl.styles.Font(bold=True, color="FFFFFF")
            cell.fill = openpyxl.styles.PatternFill(
                start_color="1E3A5F", end_color="1E3A5F", fill_type="solid")
            cell.alignment = openpyxl.styles.Alignment(horizontal="center")

        for d in sorted(summary.keys(), reverse=True):
            s = summary[d]
            pct = round((s["present"] / s["total"]) * 100, 1) if s["total"] else 0
            ws_sum.append([
                d, s["total"], s["present"], s["absent"],
                s["late"], s["on_leave"], f"{pct}%"
            ])

        wb.save(EXCEL_FILE)
    except Exception as e:
        print(f"[Excel Sync Warning] Could not sync to {EXCEL_FILE}: {e}")

def sync_leaves_to_excel():
    """Rebuild the LeaveApplications sheet in leave_data.xlsx from SQLite so
    the Excel file always reflects the current database state."""
    try:
        if os.path.exists(LEAVE_FILE):
            wb = openpyxl.load_workbook(LEAVE_FILE)
        else:
            wb = openpyxl.Workbook()

        # ── LeaveApplications sheet ───────────────────────────────
        if "LeaveApplications" in wb.sheetnames:
            del wb["LeaveApplications"]
        ws = wb.create_sheet("LeaveApplications", 0)

        headers = [
            "Leave ID", "Intern ID", "Intern Name", "Department",
            "Leave Type", "From Date", "To Date", "Days",
            "Reason", "Status", "Admin Remarks", "Applied On", "Reviewed On"
        ]
        ws.append(headers)

        # Style header row
        for col in range(1, len(headers) + 1):
            cell = ws.cell(row=1, column=col)
            cell.font = openpyxl.styles.Font(bold=True, color="FFFFFF")
            cell.fill = openpyxl.styles.PatternFill(
                start_color="1E3A5F", end_color="1E3A5F", fill_type="solid")
            cell.alignment = openpyxl.styles.Alignment(horizontal="center")

        # Fetch all leave records joined with user name and department
        conn = get_db_connection()
        cursor = conn.cursor()
        cursor.execute("""
            SELECT l.leave_id, l.user_id, u.name, u.department,
                   l.leave_type, l.from_date, l.to_date, l.days,
                   l.reason, l.status, l.remarks,
                   l.applied_on, l.reviewed_on
            FROM leaves l
            JOIN users u ON l.user_id = u.id
            ORDER BY l.applied_on DESC
        """)
        rows = cursor.fetchall()
        conn.close()

        DASH = "\u2014"
        for r in rows:
            remarks    = r["remarks"]    if r["remarks"]    and r["remarks"]    != DASH else DASH
            reviewed   = r["reviewed_on"] if r["reviewed_on"] and r["reviewed_on"] != DASH else DASH
            ws.append([
                r["leave_id"], r["user_id"], r["name"], r["department"] or "",
                r["leave_type"], r["from_date"], r["to_date"], r["days"],
                r["reason"], r["status"], remarks,
                r["applied_on"], reviewed
            ])

        # Auto-fit column widths
        col_widths = [10, 10, 22, 14, 15, 13, 13, 6, 35, 10, 28, 20, 20]
        for i, w in enumerate(col_widths, 1):
            ws.column_dimensions[
                openpyxl.utils.get_column_letter(i)].width = w

        wb.save(LEAVE_FILE)
    except Exception as e:
        print(f"[Excel Sync Warning] Could not sync to {LEAVE_FILE}: {e}")

def init_db():
    with sqlite3.connect(DB_FILE) as conn:
        cursor = conn.cursor()
        
        # 1. Users Table
        cursor.execute("""
        CREATE TABLE IF NOT EXISTS users (
            id TEXT PRIMARY KEY,
            name TEXT NOT NULL,
            email TEXT UNIQUE NOT NULL,
            password_hash TEXT NOT NULL,
            role TEXT NOT NULL CHECK(role IN ('admin', 'student')),
            department TEXT,
            created_at TIMESTAMP DEFAULT CURRENT_TIMESTAMP
        );
        """)
        
        # 2. Attendance Table
        cursor.execute("""
        CREATE TABLE IF NOT EXISTS attendance (
            attendance_id INTEGER PRIMARY KEY AUTOINCREMENT,
            user_id TEXT NOT NULL,
            date TEXT NOT NULL,
            check_in TEXT,
            check_out TEXT,
            internship_domain TEXT,
            working_hours REAL,
            status TEXT NOT NULL,
            location TEXT,
            face_verified INTEGER DEFAULT 0,
            FOREIGN KEY(user_id) REFERENCES users(id) ON DELETE CASCADE,
            UNIQUE(user_id, date)
        );
        """)
        
        # 3. Leaves Table
        cursor.execute("""
        CREATE TABLE IF NOT EXISTS leaves (
            leave_id TEXT PRIMARY KEY,
            user_id TEXT NOT NULL,
            leave_type TEXT NOT NULL,
            from_date TEXT NOT NULL,
            to_date TEXT NOT NULL,
            days INTEGER NOT NULL,
            reason TEXT NOT NULL,
            status TEXT DEFAULT 'Pending' CHECK(status IN ('Pending', 'Approved', 'Rejected')),
            remarks TEXT,
            applied_on TEXT NOT NULL,
            reviewed_on TEXT,
            FOREIGN KEY(user_id) REFERENCES users(id) ON DELETE CASCADE
        );
        """)
        conn.commit()
    
    # Run auto-migration
    migrate_excel_to_sqlite()

def migrate_excel_to_sqlite():
    with sqlite3.connect(DB_FILE) as conn:
        cursor = conn.cursor()
        
        # Seed default admin
        admin_pass_hash = generate_password_hash("admin123")
        cursor.execute("""
        INSERT OR IGNORE INTO users (id, name, email, password_hash, role, department)
        VALUES (?, ?, ?, ?, ?, ?)
        """, ("admin", "Administrator", "admin@attendance.com", admin_pass_hash, "admin", "Management"))
        
        # Check if users table is already populated with students
        cursor.execute("SELECT COUNT(*) FROM users WHERE role = 'student'")
        if cursor.fetchone()[0] > 0:
            return # Already migrated
        
        print("Starting auto-migration from Excel files to SQLite...")
        
        # Load students
        if os.path.exists(EXCEL_FILE):
            try:
                wb = openpyxl.load_workbook(EXCEL_FILE)
                if "Students" in wb.sheetnames:
                    ws = wb["Students"]
                    for row in ws.iter_rows(min_row=2, values_only=True):
                        if row[0]:
                            sid = str(row[0]).strip()
                            sname = str(row[1]).strip() if row[1] else sid
                            dept = str(row[2]).strip() if row[2] else ""
                            email = str(row[3]).strip() if row[3] else f"{sid.lower()}@company.com"
                            
                            # Password defaults to Intern ID
                            hashed_pass = generate_password_hash(sid)
                            
                            cursor.execute("""
                            INSERT OR IGNORE INTO users (id, name, email, password_hash, role, department)
                            VALUES (?, ?, ?, ?, ?, ?)
                            """, (sid, sname, email, hashed_pass, "student", dept))
            except Exception as e:
                print(f"Error migrating students: {e}")
        
        # Load attendance records
        if os.path.exists(EXCEL_FILE):
            try:
                wb = openpyxl.load_workbook(EXCEL_FILE)
                if "Attendance" in wb.sheetnames:
                    ws = wb["Attendance"]
                    for row in ws.iter_rows(min_row=2, values_only=True):
                        if row[0] and row[1]:
                            r_date = str(row[0]).split()[0] if isinstance(row[0], (datetime, datetime.date)) else str(row[0]).strip()
                            sid = str(row[1]).strip()
                            
                            # Skip if user doesn't exist
                            cursor.execute("SELECT 1 FROM users WHERE id = ?", (sid,))
                            if not cursor.fetchone():
                                hashed_pass = generate_password_hash(sid)
                                cursor.execute("""
                                INSERT OR IGNORE INTO users (id, name, email, password_hash, role, department)
                                VALUES (?, ?, ?, ?, ?, ?)
                                """, (sid, str(row[2]).strip() if row[2] else sid, f"{sid.lower()}@company.com", hashed_pass, "student", ""))
                            
                            login_t = str(row[3]).strip() if row[3] else None
                            if login_t and len(login_t) > 8: # Handle datetime objects
                                login_t = login_t.split()[-1]
                            
                            logout_t = str(row[4]).strip() if row[4] else None
                            if logout_t and len(logout_t) > 8:
                                logout_t = logout_t.split()[-1]
                            
                            hrs_str = str(row[5]).strip() if row[5] else "0"
                            hrs = 0.0
                            try:
                                hrs = float(hrs_str.replace(" hrs", "").replace("—", "0"))
                            except:
                                pass
                            
                            status = str(row[6]).strip() if row[6] else "Present"
                            remarks = str(row[7]).strip() if row[7] else ""
                            domain = "Other"
                            
                            cursor.execute("""
                            INSERT OR REPLACE INTO attendance (user_id, date, check_in, check_out, internship_domain, working_hours, status, location, face_verified)
                            VALUES (?, ?, ?, ?, ?, ?, ?, ?, ?)
                            """, (sid, r_date, login_t, logout_t, domain, hrs, status, remarks, 0))
            except Exception as e:
                print(f"Error migrating attendance: {e}")

        # Load leaves
        if os.path.exists(LEAVE_FILE):
            try:
                wb = openpyxl.load_workbook(LEAVE_FILE)
                if "LeaveApplications" in wb.sheetnames:
                    ws = wb["LeaveApplications"]
                    for row in ws.iter_rows(min_row=2, values_only=True):
                        if row[0] and row[1]:
                            lid = str(row[0]).strip()
                            sid = str(row[1]).strip()
                            
                            cursor.execute("SELECT 1 FROM users WHERE id = ?", (sid,))
                            if not cursor.fetchone():
                                continue
                            
                            ltype = str(row[4]).strip() if row[4] else "Other"
                            from_d = str(row[5]).strip() if row[5] else ""
                            to_d = str(row[6]).strip() if row[6] else ""
                            
                            days = 1
                            try:
                                days = int(row[7])
                            except:
                                pass
                            
                            reason = str(row[8]).strip() if row[8] else ""
                            status = str(row[9]).strip() if row[9] else "Pending"
                            remarks = str(row[10]).strip() if row[10] else ""
                            applied = str(row[11]).strip() if row[11] else ""
                            reviewed = str(row[12]).strip() if row[12] else ""
                            
                            cursor.execute("""
                            INSERT OR IGNORE INTO leaves (leave_id, user_id, leave_type, from_date, to_date, days, reason, status, remarks, applied_on, reviewed_on)
                            VALUES (?, ?, ?, ?, ?, ?, ?, ?, ?, ?, ?)
                            """, (lid, sid, ltype, from_d, to_d, days, reason, status, remarks, applied, reviewed))
            except Exception as e:
                print(f"Error migrating leaves: {e}")
        
        conn.commit()
        print("Auto-migration to SQLite finished successfully!")

# ─────────────────────────────────────────────────────────────
#  SECURITY ROUTE MIDDLEWARE
# ─────────────────────────────────────────────────────────────
def login_required(f):
    @wraps(f)
    def decorated_function(*args, **kwargs):
        if 'user_id' not in session:
            if request.path.startswith('/api/'):
                return jsonify(success=False, error="Unauthorized", message="Authentication required."), 401
            return redirect(url_for('login_page'))
        return f(*args, **kwargs)
    return decorated_function

def roles_accepted(*roles):
    def decorator(f):
        @wraps(f)
        def decorated_function(*args, **kwargs):
            if session.get('role') not in roles:
                if request.path.startswith('/api/'):
                    return jsonify(success=False, error="Forbidden", message="Access denied."), 403
                return render_template('error.html', code=403, message="Access denied. You do not have permission to access this page."), 403
            return f(*args, **kwargs)
        return decorated_function
    return decorator

# ─────────────────────────────────────────────────────────────
#  PAGE ROUTING
# ─────────────────────────────────────────────────────────────
@app.route("/")
def login_page():
    if 'user_id' in session:
        if session.get('role') == 'admin':
            return redirect(url_for('admin_dashboard'))
        else:
            return redirect(url_for('student_dashboard'))
    return render_template("login.html")

@app.route("/login")
def login_redirect():
    return redirect(url_for('login_page'))

@app.route("/student/dashboard")
@login_required
@roles_accepted('student')
def student_dashboard():
    return render_template("student_dashboard.html")

@app.route("/admin")
@app.route("/admin/dashboard")
@login_required
@roles_accepted('admin')
def admin_dashboard():
    return render_template("admin_dashboard.html")

@app.route("/admin/users")
@login_required
@roles_accepted('admin')
def admin_users():
    return redirect(url_for('admin_dashboard') + '#interns')

@app.route("/admin/reports")
@login_required
@roles_accepted('admin')
def admin_reports():
    return redirect(url_for('admin_dashboard') + '#records')

# ─────────────────────────────────────────────────────────────
#  API — AUTHENTICATION
# ─────────────────────────────────────────────────────────────
@app.route("/api/auth/login", methods=["POST"])
def auth_login():
    # Rate Limiting
    ip = request.remote_addr
    if login_limiter.is_blocked(ip):
        return jsonify(success=False, message="Too many login attempts. Please try again in a minute."), 429

    d = request.get_json() or {}
    username = d.get("username", "").strip()
    password = d.get("password", "").strip()
    role = d.get("role", "student").strip()

    if not username or not password or not role:
        return jsonify(success=False, message="Username, Password, and Role are required."), 400
    if len(username) > 100 or len(password) > 100 or len(role) > 50:
        return jsonify(success=False, message="Inputs exceed maximum allowed length."), 400
    if role not in ('admin', 'student'):
        return jsonify(success=False, message="Invalid role specified."), 400

    conn = get_db_connection()
    cursor = conn.cursor()
    # Check if credentials are correct
    cursor.execute("SELECT * FROM users WHERE id = ? OR email = ?", (username, username))
    user = cursor.fetchone()
    conn.close()

    if user and check_password_hash(user["password_hash"], password):
        if user["role"] != role:
            return jsonify(success=False, message=f"Incorrect portal. User is registered as a {user['role']}.")
        
        # Save CSRF token before clearing session
        csrf = session.get("csrf_token")
        session.clear()
        if csrf:
            session["csrf_token"] = csrf
        else:
            session["csrf_token"] = secrets.token_hex(32)

        session["user_id"] = user["id"]
        session["role"] = user["role"]
        session["name"] = user["name"]
        session["email"] = user["email"]
        session["department"] = user["department"]
        
        return jsonify(success=True, role=user["role"], name=user["name"])
    
    return jsonify(success=False, message="Invalid username, email, or password.")

@app.route("/api/auth/logout", methods=["POST"])
def auth_logout():
    # Save CSRF token before clearing
    csrf = session.get("csrf_token")
    session.clear()
    if csrf:
        session["csrf_token"] = csrf
    else:
        session["csrf_token"] = secrets.token_hex(32)
    return jsonify(success=True, message="Session ended.")

def validate_password_strength(password):
    if len(password) < 8:
        return False, "Password must be at least 8 characters long."
    if not re.search(r"[A-Z]", password):
        return False, "Password must contain at least one uppercase letter."
    if not re.search(r"[a-z]", password):
        return False, "Password must contain at least one lowercase letter."
    if not re.search(r"\d", password):
        return False, "Password must contain at least one number."
    if not re.search(r"[!@#$%^&*(),.?\":{}|<>]", password):
        return False, "Password must contain at least one special character."
    return True, ""

@app.route("/change-password", methods=["GET", "POST"])
@login_required
def change_password_page():
    if request.method == "POST":
        current_password = request.form.get("current_password", "").strip()
        new_password = request.form.get("new_password", "").strip()
        confirm_password = request.form.get("confirm_password", "").strip()

        if not current_password or not new_password or not confirm_password:
            flash("All fields are required.", "error")
            return redirect(url_for("change_password_page"))

        if len(current_password) > 100 or len(new_password) > 100 or len(confirm_password) > 100:
            flash("Input values exceed maximum allowed length.", "error")
            return redirect(url_for("change_password_page"))

        if new_password != confirm_password:
            flash("New password and confirm password do not match.", "error")
            return redirect(url_for("change_password_page"))

        is_strong, err_msg = validate_password_strength(new_password)
        if not is_strong:
            flash(err_msg, "error")
            return redirect(url_for("change_password_page"))

        user_id = session.get("user_id")
        conn = get_db_connection()
        cursor = conn.cursor()
        cursor.execute("SELECT password_hash FROM users WHERE id = ?", (user_id,))
        user = cursor.fetchone()
        
        if not user or not check_password_hash(user["password_hash"], current_password):
            conn.close()
            flash("Incorrect current password.", "error")
            return redirect(url_for("change_password_page"))

        new_hash = generate_password_hash(new_password)
        cursor.execute("UPDATE users SET password_hash = ? WHERE id = ?", (new_hash, user_id))
        conn.commit()
        conn.close()

        flash("Password updated successfully!", "success")
        return redirect(url_for("change_password_page"))

    return render_template("change_password.html")

@app.route("/api/auth/me")
@login_required
def auth_me():
    return jsonify(
        logged_in=True,
        id=session.get("user_id"),
        name=session.get("name"),
        email=session.get("email"),
        role=session.get("role"),
        department=session.get("department")
    )

# ─────────────────────────────────────────────────────────────
#  API — STUDENT INTERFACES
# ─────────────────────────────────────────────────────────────
@app.route("/api/student/check-in", methods=["POST"])
@login_required
@roles_accepted('student')
def student_check_in():
    d = request.get_json() or {}
    domain = d.get("internship_domain", "").strip()
    location = d.get("location", "Browser Client").strip()
    
    if not domain:
        return jsonify(success=False, message="Internship domain selection is required."), 400
    if len(domain) > 100 or len(location) > 200:
        return jsonify(success=False, message="Inputs exceed maximum allowed length."), 400

    domain = html.escape(domain)
    location = html.escape(location)

    user_id = session.get("user_id")
    now = datetime.now()
    date_str = now.strftime("%Y-%m-%d")
    time_str = now.strftime("%H:%M:%S")

    # Cutoff for Late
    cutoff = now.replace(hour=LOGIN_HOUR, minute=LATE_MIN, second=0, microsecond=0)
    status = "Late" if now > cutoff else "Logged In"

    conn = get_db_connection()
    cursor = conn.cursor()
    
    # Check if leave is approved for today
    cursor.execute("SELECT 1 FROM leaves WHERE user_id = ? AND ? BETWEEN from_date AND to_date AND status = 'Approved'", (user_id, date_str))
    if cursor.fetchone():
        conn.close()
        return jsonify(success=False, message="You are marked On Leave for today. Check-in not allowed.")

    # Check if already checked in today
    cursor.execute("SELECT * FROM attendance WHERE user_id = ? AND date = ?", (user_id, date_str))
    record = cursor.fetchone()
    
    if record:
        conn.close()
        return jsonify(success=False, message="You have already checked in today.")

    try:
        cursor.execute("""
        INSERT INTO attendance (user_id, date, check_in, check_out, internship_domain, working_hours, status, location, face_verified)
        VALUES (?, ?, ?, ?, ?, ?, ?, ?, ?)
        """, (user_id, date_str, time_str, "—", domain, 0.0, status, location, 0))
        conn.commit()
        success = True
        msg = f"Check-In recorded at {time_str} ({status})"
    except Exception as e:
        success = False
        msg = f"Database error: {str(e)}"
    
    conn.close()

    # Sync the updated attendance back to Excel
    if success:
        sync_attendance_to_excel()

    return jsonify(success=success, message=msg)

@app.route("/api/student/check-out", methods=["POST"])
@login_required
@roles_accepted('student')
def student_check_out():
    user_id = session.get("user_id")
    now = datetime.now()
    date_str = now.strftime("%Y-%m-%d")
    time_str = now.strftime("%H:%M:%S")

    conn = get_db_connection()
    cursor = conn.cursor()

    cursor.execute("SELECT * FROM attendance WHERE user_id = ? AND date = ?", (user_id, date_str))
    record = cursor.fetchone()

    if not record:
        conn.close()
        return jsonify(success=False, message="No Check-In record found for today.")
    
    if record["check_out"] and record["check_out"] != "—":
        conn.close()
        return jsonify(success=False, message="You have already checked out today.")

    login_t_str = record["check_in"]
    hrs = 0.0
    try:
        lt = datetime.strptime(f"{date_str} {login_t_str}", "%Y-%m-%d %H:%M:%S")
        lo = datetime.strptime(f"{date_str} {time_str}", "%Y-%m-%d %H:%M:%S")
        hrs = round((lo - lt).total_seconds() / 3600, 2)
    except Exception as e:
        print("Hours calculation error:", e)

    try:
        cursor.execute("""
        UPDATE attendance 
        SET check_out = ?, working_hours = ?, status = 'Present'
        WHERE user_id = ? AND date = ?
        """, (time_str, hrs, user_id, date_str))
        conn.commit()
        success = True
        msg = f"Check-Out recorded at {time_str}. Total hours: {hrs:.2f} hrs."
    except Exception as e:
        success = False
        msg = f"Database update error: {str(e)}"

    conn.close()

    # Sync the updated attendance back to Excel
    if success:
        sync_attendance_to_excel()

    return jsonify(success=success, message=msg)

@app.route("/api/student/attendance")
@login_required
@roles_accepted('student')
def student_attendance():
    user_id = session.get("user_id")
    conn = get_db_connection()
    cursor = conn.cursor()
    cursor.execute("""
    SELECT date, check_in, check_out, internship_domain, working_hours, status, location
    FROM attendance
    WHERE user_id = ?
    ORDER BY date DESC
    """, (user_id,))
    rows = cursor.fetchall()
    conn.close()

    records = [dict(r) for r in rows]
    return jsonify(success=True, records=records)

@app.route("/api/student/leave/apply", methods=["POST"])
@login_required
@roles_accepted('student')
def student_leave_apply():
    d = request.get_json() or {}
    ltype = d.get("leave_type", "").strip()
    from_date = d.get("from_date", "").strip()
    to_date = d.get("to_date", "").strip()
    reason = d.get("reason", "").strip()

    if not all([ltype, from_date, to_date, reason]):
        return jsonify(success=False, message="All fields are required."), 400
    if len(ltype) > 100 or len(from_date) > 20 or len(to_date) > 20 or len(reason) > 1000:
        return jsonify(success=False, message="Inputs exceed maximum allowed length."), 400
    if not is_valid_date(from_date) or not is_valid_date(to_date):
        return jsonify(success=False, message="Invalid date format. Expected YYYY-MM-DD."), 400

    # Escape inputs
    ltype = html.escape(ltype)
    reason = html.escape(reason)

    user_id = session.get("user_id")
    
    try:
        f = datetime.strptime(from_date, "%Y-%m-%d")
        t = datetime.strptime(to_date, "%Y-%m-%d")
        days = max(1, (t - f).days + 1)
    except:
        return jsonify(success=False, message="Invalid date calculations."), 400

    conn = get_db_connection()
    cursor = conn.cursor()

    # Generate next Leave ID
    cursor.execute("SELECT leave_id FROM leaves ORDER BY leave_id DESC LIMIT 1")
    last_row = cursor.fetchone()
    next_num = 1
    if last_row:
        try:
            next_num = int(last_row["leave_id"].replace("LV", "")) + 1
        except:
            pass
    lid = f"LV{next_num:04d}"

    applied_on = datetime.now().strftime("%Y-%m-%d %H:%M:%S")

    try:
        cursor.execute("""
        INSERT INTO leaves (leave_id, user_id, leave_type, from_date, to_date, days, reason, status, remarks, applied_on, reviewed_on)
        VALUES (?, ?, ?, ?, ?, ?, ?, ?, ?, ?, ?)
        """, (lid, user_id, ltype, from_date, to_date, days, reason, "Pending", "—", applied_on, "—"))
        conn.commit()
        success = True
        msg = f"Leave application {lid} submitted successfully."
    except Exception as e:
        success = False
        msg = f"Database error: {str(e)}"

    conn.close()

    # Sync the updated leaves back to Excel
    if success:
        sync_leaves_to_excel()

    return jsonify(success=success, message=msg)

@app.route("/api/student/leave/my_leaves")
@login_required
@roles_accepted('student')
def student_my_leaves():
    user_id = session.get("user_id")
    conn = get_db_connection()
    cursor = conn.cursor()
    cursor.execute("""
    SELECT leave_id, leave_type, from_date, to_date, days, reason, status, remarks, applied_on
    FROM leaves
    WHERE user_id = ?
    ORDER BY applied_on DESC
    """, (user_id,))
    rows = cursor.fetchall()
    conn.close()

    leaves = [dict(r) for r in rows]
    return jsonify(success=True, leaves=leaves)

# ─────────────────────────────────────────────────────────────
#  API — ADMIN INTERFACES
# ─────────────────────────────────────────────────────────────
def build_admin_filter_query(request_args):
    where_clauses = ["u.role = ?"]
    params = ["student"]

    domain = request_args.get("domain")
    if domain:
        domain = domain.strip()
        if len(domain) <= 100:
            where_clauses.append("a.internship_domain = ?")
            params.append(domain)

    status = request_args.get("status")
    if status:
        status = status.strip()
        if len(status) <= 100:
            where_clauses.append("a.status = ?")
            params.append(status)

    name = request_args.get("name")
    if name:
        name = name.strip()
        if len(name) <= 100:
            where_clauses.append("u.name LIKE ?")
            params.append(f"%{name}%")

    sid = request_args.get("id")
    if sid:
        sid = sid.strip()
        if len(sid) <= 50:
            where_clauses.append("a.user_id LIKE ?")
            params.append(f"%{sid}%")

    date_range = request_args.get("date_range", "today")
    today = datetime.now()
    
    if date_range == "today":
        where_clauses.append("a.date = ?")
        params.append(today.strftime("%Y-%m-%d"))
    elif date_range == "yesterday":
        yst = (today - timedelta(days=1)).strftime("%Y-%m-%d")
        where_clauses.append("a.date = ?")
        params.append(yst)
    elif date_range == "7days":
        cutoff = (today - timedelta(days=7)).strftime("%Y-%m-%d")
        where_clauses.append("a.date >= ?")
        params.append(cutoff)
    elif date_range == "30days":
        cutoff = (today - timedelta(days=30)).strftime("%Y-%m-%d")
        where_clauses.append("a.date >= ?")
        params.append(cutoff)
    elif date_range == "custom":
        start_date = request_args.get("start_date")
        end_date = request_args.get("end_date")
        if start_date and is_valid_date(start_date.strip()):
            where_clauses.append("a.date >= ?")
            params.append(start_date.strip())
        if end_date and is_valid_date(end_date.strip()):
            where_clauses.append("a.date <= ?")
            params.append(end_date.strip())
            
    return where_clauses, params

@app.route("/api/admin/stats")
@login_required
@roles_accepted('admin')
def admin_stats():
    where_clauses, params = build_admin_filter_query(request.args)
    where_str = " AND ".join(where_clauses)

    conn = get_db_connection()
    cursor = conn.cursor()

    # 1. KPIs
    # Registered students
    name_filter = request.args.get("name")
    id_filter = request.args.get("id")
    reg_where = ["role = 'student'"]
    reg_params = []
    if name_filter:
        name_filter = name_filter.strip()
        if len(name_filter) <= 100:
            reg_where.append("name LIKE ?")
            reg_params.append(f"%{name_filter}%")
    if id_filter:
        id_filter = id_filter.strip()
        if len(id_filter) <= 50:
            reg_where.append("id LIKE ?")
            reg_params.append(f"%{id_filter}%")
    
    cursor.execute(f"SELECT COUNT(*) FROM users WHERE {' AND '.join(reg_where)}", reg_params)
    total_users = cursor.fetchone()[0]

    # Present Today
    cursor.execute(f"SELECT COUNT(*) FROM attendance a JOIN users u ON a.user_id = u.id WHERE {where_str} AND a.status IN ('Present', 'Logged In')", params)
    present_today = cursor.fetchone()[0]

    # Absent Today
    cursor.execute(f"SELECT COUNT(*) FROM attendance a JOIN users u ON a.user_id = u.id WHERE {where_str} AND a.status = 'Absent'", params)
    absent_today = cursor.fetchone()[0]

    # Late arrivals
    cursor.execute(f"SELECT COUNT(*) FROM attendance a JOIN users u ON a.user_id = u.id WHERE {where_str} AND a.status = 'Late'", params)
    late_today = cursor.fetchone()[0]

    # Avg Working Hours
    cursor.execute(f"SELECT AVG(working_hours) FROM attendance a JOIN users u ON a.user_id = u.id WHERE {where_str} AND working_hours > 0", params)
    avg_hours = cursor.fetchone()[0] or 0.0

    # 2. Charts Data
    # Domain-wise distribution
    cursor.execute(f"""
    SELECT a.internship_domain, COUNT(*) 
    FROM attendance a JOIN users u ON a.user_id = u.id 
    WHERE {where_str} 
    GROUP BY a.internship_domain
    """, params)
    domains_chart = {row[0]: row[1] for row in cursor.fetchall()}

    # Trends (last 7 days counts)
    trends_where, trends_params = [], []
    # copy filter parameters except date_range
    for c, p in zip(where_clauses, params):
        if "a.date" not in c:
            trends_where.append(c)
            trends_params.append(p)
            
    trends_cutoff = (datetime.now() - timedelta(days=7)).strftime("%Y-%m-%d")
    trends_where.append("a.date >= ?")
    trends_params.append(trends_cutoff)
    trends_where_str = " AND ".join(trends_where)

    cursor.execute(f"""
    SELECT a.date, a.status, COUNT(*) 
    FROM attendance a JOIN users u ON a.user_id = u.id 
    WHERE {trends_where_str} 
    GROUP BY a.date, a.status
    """, trends_params)
    
    trends_chart = {}
    for row in cursor.fetchall():
        r_date, r_status, r_count = row[0], row[1], row[2]
        if r_date not in trends_chart:
            trends_chart[r_date] = {"present": 0, "absent": 0, "late": 0}
        if r_status in ("Present", "Logged In"):
            trends_chart[r_date]["present"] += r_count
        elif r_status == "Absent":
            trends_chart[r_date]["absent"] += r_count
        elif r_status == "Late":
            trends_chart[r_date]["late"] += r_count

    # Daily check-in timeline (frequency by hour)
    cursor.execute(f"""
    SELECT SUBSTR(a.check_in, 1, 2) AS hr, COUNT(*)
    FROM attendance a JOIN users u ON a.user_id = u.id
    WHERE {where_str} AND a.check_in IS NOT NULL AND a.check_in != '—'
    GROUP BY hr
    """, params)
    timeline_chart = {f"{row[0]}:00": row[1] for row in cursor.fetchall()}

    conn.close()

    return jsonify(
        success=True,
        stats={
            "total_users": total_users,
            "present_today": present_today,
            "absent_today": absent_today,
            "late_today": late_today,
            "avg_hours": avg_hours
        },
        charts={
            "domains": domains_chart,
            "trends": trends_chart,
            "timeline": timeline_chart
        }
    )

@app.route("/api/admin/attendance")
@login_required
@roles_accepted('admin')
def admin_attendance():
    where_clauses, params = build_admin_filter_query(request.args)
    where_str = " AND ".join(where_clauses)

    conn = get_db_connection()
    cursor = conn.cursor()
    cursor.execute(f"""
    SELECT a.date, a.user_id, u.name, a.internship_domain, a.check_in, a.check_out, a.working_hours, a.status
    FROM attendance a
    JOIN users u ON a.user_id = u.id
    WHERE {where_str}
    ORDER BY a.date DESC, a.check_in DESC
    """, params)
    rows = cursor.fetchall()
    conn.close()

    records = [dict(r) for r in rows]
    return jsonify(success=True, records=records)

@app.route("/api/admin/interns")
@login_required
@roles_accepted('admin')
def admin_interns():
    conn = get_db_connection()
    cursor = conn.cursor()
    cursor.execute("SELECT id, name, email, department, created_at FROM users WHERE role = 'student' ORDER BY id ASC")
    rows = cursor.fetchall()
    conn.close()

    interns = [dict(r) for r in rows]
    return jsonify(success=True, interns=interns)

@app.route("/api/admin/interns/add", methods=["POST"])
@login_required
@roles_accepted('admin')
def admin_add_intern():
    d = request.get_json() or {}
    sid = d.get("intern_id", "").strip()
    name = d.get("name", "").strip()
    email = d.get("email", "").strip()
    dept = d.get("department", "").strip()

    if not sid or not name or not email:
        return jsonify(success=False, message="ID, Name, and Email are required."), 400
    if len(sid) > 50 or len(name) > 100 or len(email) > 100 or len(dept) > 100:
        return jsonify(success=False, message="Inputs exceed maximum allowed length."), 400
    if not is_alphanumeric(sid):
        return jsonify(success=False, message="Intern ID must be alphanumeric."), 400
    if not is_valid_email(email):
        return jsonify(success=False, message="Invalid email address format."), 400

    # Escape inputs
    sid = html.escape(sid)
    name = html.escape(name)
    email = html.escape(email)
    dept = html.escape(dept)

    # Password defaults to Intern ID itself
    hashed_pass = generate_password_hash(sid)

    conn = get_db_connection()
    cursor = conn.cursor()
    try:
        cursor.execute("""
        INSERT INTO users (id, name, email, password_hash, role, department)
        VALUES (?, ?, ?, ?, 'student', ?)
        """, (sid, name, email, hashed_pass, dept))
        conn.commit()
        success = True
        msg = f"Student {name} registered successfully."
    except sqlite3.IntegrityError:
        success = False
        msg = "Intern ID or Email already registered."
    except Exception as e:
        success = False
        msg = f"Database error: {str(e)}"
    
    conn.close()
    return jsonify(success=success, message=msg)

@app.route("/api/admin/interns/<sid>", methods=["DELETE"])
@login_required
@roles_accepted('admin')
def admin_delete_intern(sid):
    if not sid or len(sid) > 50 or not is_alphanumeric(sid):
        return jsonify(success=False, message="Invalid Student ID format."), 400

    sid = html.escape(sid)
    conn = get_db_connection()
    cursor = conn.cursor()
    cursor.execute("SELECT 1 FROM users WHERE id = ? AND role = 'student'", (sid,))
    if not cursor.fetchone():
        conn.close()
        return jsonify(success=False, message="Student record not found.")

    try:
        cursor.execute("DELETE FROM users WHERE id = ?", (sid,))
        conn.commit()
        success = True
        msg = "Student record deleted successfully."
    except Exception as e:
        success = False
        msg = f"Database error: {str(e)}"
    
    conn.close()
    return jsonify(success=success, message=msg)

@app.route("/api/admin/leaves")
@login_required
@roles_accepted('admin')
def admin_leaves():
    conn = get_db_connection()
    cursor = conn.cursor()
    cursor.execute("""
    SELECT l.leave_id, l.user_id, u.name, l.leave_type, l.from_date, l.to_date, l.days, l.reason, l.status, l.remarks, l.reviewed_on
    FROM leaves l
    JOIN users u ON l.user_id = u.id
    ORDER BY l.applied_on DESC
    """)
    rows = cursor.fetchall()
    conn.close()

    leaves = [dict(r) for r in rows]
    return jsonify(success=True, leaves=leaves)

@app.route("/api/admin/leave/review", methods=["POST"])
@login_required
@roles_accepted('admin')
def admin_leave_review():
    d = request.get_json() or {}
    lid = d.get("leave_id", "").strip()
    action = d.get("action", "").strip() # 'Approved' / 'Rejected'
    remarks = d.get("remarks", "").strip()

    if not lid or not action:
        return jsonify(success=False, message="Leave ID and Action are required."), 400
    if action not in ('Approved', 'Rejected'):
        return jsonify(success=False, message="Invalid review action."), 400
    if len(lid) > 50 or len(remarks) > 1000:
        return jsonify(success=False, message="Inputs exceed maximum allowed length."), 400
    if not is_alphanumeric(lid.replace("LV", "")):
        return jsonify(success=False, message="Invalid Leave ID format."), 400

    # Escape inputs
    lid = html.escape(lid)
    remarks = html.escape(remarks)

    reviewed_on = datetime.now().strftime("%Y-%m-%d %H:%M:%S")

    conn = get_db_connection()
    cursor = conn.cursor()
    
    cursor.execute("SELECT * FROM leaves WHERE leave_id = ?", (lid,))
    leave = cursor.fetchone()
    if not leave:
        conn.close()
        return jsonify(success=False, message="Leave application not found.")

    if leave["status"] != "Pending":
        conn.close()
        return jsonify(success=False, message="This application has already been reviewed.")

    try:
        cursor.execute("""
        UPDATE leaves 
        SET status = ?, remarks = ?, reviewed_on = ?
        WHERE leave_id = ?
        """, (action, remarks or "Reviewed", reviewed_on, lid))

        # If approved, we need to auto-insert "On Leave" attendance records for dates covered
        if action == "Approved":
            user_id = leave["user_id"]
            f_date = datetime.strptime(leave["from_date"], "%Y-%m-%d")
            t_date = datetime.strptime(leave["to_date"], "%Y-%m-%d")
            curr = f_date
            while curr <= t_date:
                curr_str = curr.strftime("%Y-%m-%d")
                # Insert or replace attendance as On Leave
                cursor.execute("""
                INSERT OR REPLACE INTO attendance (user_id, date, check_in, check_out, internship_domain, working_hours, status, location, face_verified)
                VALUES (?, ?, '—', '—', 'Leave', 0.0, 'On Leave', ?, 0)
                """, (user_id, curr_str, f"Approved Leave {lid}"))
                curr += timedelta(days=1)

        conn.commit()
        success = True
        msg = f"Leave application {lid} has been {action.lower()}."
    except Exception as e:
        success = False
        msg = f"Error updating database: {str(e)}"

    conn.close()

    # Sync the updated leaves (and attendance for approved leave dates) back to Excel
    if success:
        sync_leaves_to_excel()
        sync_attendance_to_excel()

    return jsonify(success=success, message=msg)

@app.route("/api/admin/download_excel")
@login_required
@roles_accepted('admin')
def admin_download_excel():
    where_clauses, params = build_admin_filter_query(request.args)
    where_str = " AND ".join(where_clauses)

    conn = get_db_connection()
    cursor = conn.cursor()
    cursor.execute(f"""
    SELECT a.date, a.user_id, u.name, a.internship_domain, a.check_in, a.check_out, a.working_hours, a.status
    FROM attendance a
    JOIN users u ON a.user_id = u.id
    WHERE {where_str}
    ORDER BY a.date DESC
    """, params)
    rows = cursor.fetchall()
    conn.close()

    # Generate workbook on the fly
    wb = openpyxl.Workbook()
    ws = wb.active
    ws.title = "Attendance Logs"
    
    headers = ["Date", "Intern ID", "Intern Name", "Internship Domain", "Check-In", "Check-Out", "Working Hours", "Status"]
    ws.append(headers)
    
    # Header styles
    for col_num in range(1, 9):
        cell = ws.cell(row=1, column=col_num)
        cell.font = openpyxl.styles.Font(bold=True, color="FFFFFF")
        cell.fill = openpyxl.styles.PatternFill(start_color="1E3A5F", end_color="1E3A5F", fill_type="solid")
        cell.alignment = openpyxl.styles.Alignment(horizontal="center")
    
    for r in rows:
        ws.append([
            r["date"],
            r["user_id"],
            r["name"],
            r["internship_domain"] or "—",
            r["check_in"] or "—",
            r["check_out"] or "—",
            f"{r['working_hours']:.2f} hrs" if r["working_hours"] else "0.00 hrs",
            r["status"]
        ])
        
    # Set column widths
    for col in ws.columns:
        max_len = max(len(str(cell.value or '')) for cell in col)
        col_letter = openpyxl.utils.get_column_letter(col[0].column)
        ws.column_dimensions[col_letter].width = max(max_len + 3, 12)

    out = io.BytesIO()
    wb.save(out)
    out.seek(0)
    
    return send_file(
        out,
        as_attachment=True,
        download_name=f"attendance_report_{datetime.now().strftime('%Y%m%d')}.xlsx",
        mimetype="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet"
    )

@app.route("/api/admin/export_powerbi")
@login_required
@roles_accepted('admin')
def admin_export_powerbi():
    where_clauses, params = build_admin_filter_query(request.args)
    where_str = " AND ".join(where_clauses)

    conn = get_db_connection()
    cursor = conn.cursor()
    cursor.execute(f"""
    SELECT a.date, a.user_id, u.name, a.internship_domain, a.check_in, a.check_out, a.working_hours, a.status
    FROM attendance a
    JOIN users u ON a.user_id = u.id
    WHERE {where_str}
    ORDER BY a.date DESC
    """, params)
    rows = cursor.fetchall()
    conn.close()

    # Create CSV Response
    import csv
    output = io.StringIO()
    writer = csv.writer(output)
    writer.writerow(["Date", "InternID", "InternName", "InternshipDomain", "LoginTime", "LogoutTime", "WorkingHours", "Status"])
    
    for r in rows:
        writer.writerow([
            r["date"],
            r["user_id"],
            r["name"],
            r["internship_domain"] or "—",
            r["check_in"] or "—",
            r["check_out"] or "—",
            f"{r['working_hours']:.2f}" if r["working_hours"] else "0.00",
            r["status"]
        ])
    
    output.seek(0)
    return Response(
        output.getvalue(),
        mimetype="text/csv",
        headers={"Content-Disposition": "attachment; filename=attendance_powerbi.csv"}
    )

# ─────────────────────────────────────────────────────────────
#  MAIN SERVER START
# ─────────────────────────────────────────────────────────────
if __name__ == "__main__":
    init_db()
    print("\n====================================================")
    print("  AI-Based Attendance System (SQLite & RBAC Mode)")
    print("  http://localhost:5000")
    print("----------------------------------------------------")
    print("  Admin Credentials: admin / admin123")
    print("  Student Default Passwords: [their Student/Intern ID]")
    print("====================================================\n")
    app.run(debug=False, port=5000)