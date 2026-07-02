import openpyxl, os, sqlite3

DB_FILE = 'attendance.db'
LEAVE_FILE = 'leave_data.xlsx'
DASH = '\u2014'

conn = sqlite3.connect(DB_FILE)
conn.row_factory = sqlite3.Row
cursor = conn.cursor()
cursor.execute("""
    SELECT l.leave_id, l.user_id, u.name, u.department,
           l.leave_type, l.from_date, l.to_date, l.days,
           l.reason, l.status, l.remarks,
           l.applied_on, l.reviewed_on
    FROM leaves l
    JOIN users u ON l.user_id = u.id
    ORDER BY l.applied_on DESC
""")
rows = cursor.fetchall()
conn.close()

if os.path.exists(LEAVE_FILE):
    wb = openpyxl.load_workbook(LEAVE_FILE)
else:
    wb = openpyxl.Workbook()

if 'LeaveApplications' in wb.sheetnames:
    del wb['LeaveApplications']
ws = wb.create_sheet('LeaveApplications', 0)

headers = [
    'Leave ID', 'Intern ID', 'Intern Name', 'Department',
    'Leave Type', 'From Date', 'To Date', 'Days',
    'Reason', 'Status', 'Admin Remarks', 'Applied On', 'Reviewed On'
]
ws.append(headers)
for col in range(1, len(headers)+1):
    cell = ws.cell(row=1, column=col)
    cell.font = openpyxl.styles.Font(bold=True, color='FFFFFF')
    cell.fill = openpyxl.styles.PatternFill(start_color='1E3A5F', end_color='1E3A5F', fill_type='solid')
    cell.alignment = openpyxl.styles.Alignment(horizontal='center')

for r in rows:
    remarks  = r['remarks']    if r['remarks']    and r['remarks']    != DASH else DASH
    reviewed = r['reviewed_on'] if r['reviewed_on'] and r['reviewed_on'] != DASH else DASH
    ws.append([
        r['leave_id'], r['user_id'], r['name'], r['department'] or '',
        r['leave_type'], r['from_date'], r['to_date'], r['days'],
        r['reason'], r['status'], remarks,
        r['applied_on'], reviewed
    ])

col_widths = [10, 10, 22, 14, 15, 13, 13, 6, 35, 10, 28, 20, 20]
for i, w in enumerate(col_widths, 1):
    ws.column_dimensions[openpyxl.utils.get_column_letter(i)].width = w

wb.save(LEAVE_FILE)
print(f'Done! Rows written to LeaveApplications sheet: {ws.max_row - 1}')
for r in rows:
    print(f"  {r['leave_id']} | {r['user_id']} | {r['name']} | {r['status']}")
