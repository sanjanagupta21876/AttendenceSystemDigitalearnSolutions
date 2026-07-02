import openpyxl

wb = openpyxl.load_workbook('attendance_data.xlsx')
print('Sheets:', wb.sheetnames)
for s in wb.sheetnames:
    ws = wb[s]
    print(f'\n--- Sheet: {s} (rows: {ws.max_row}) ---')
    for i, row in enumerate(ws.iter_rows(values_only=True), 1):
        print(f'  Row {i}: {list(row)}')
        if i > 10:
            print('  ... (truncated)')
            break
