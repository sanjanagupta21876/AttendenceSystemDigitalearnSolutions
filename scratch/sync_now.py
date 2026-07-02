import openpyxl, os, sqlite3

DB_FILE = 'attendance.db'
EXCEL_FILE = 'attendance_data.xlsx'

conn = sqlite3.connect(DB_FILE)
conn.row_factory = sqlite3.Row
cursor = conn.cursor()
cursor.execute("""
    SELECT a.date, a.user_id, u.name,
           a.check_in, a.check_out,
           a.working_hours, a.status, a.location
    FROM attendance a
    JOIN users u ON a.user_id = u.id
    ORDER BY a.date DESC, a.user_id
""")
rows = cursor.fetchall()
conn.close()

wb = openpyxl.load_workbook(EXCEL_FILE)

# Rebuild Attendance sheet
if 'Attendance' in wb.sheetnames:
    del wb['Attendance']
ws_att = wb.create_sheet('Attendance', 0)

headers = ['Date','Intern ID','Intern Name','Login Time','Logout Time','Working Hours','Status','Remarks']
ws_att.append(headers)
for col in range(1, len(headers)+1):
    cell = ws_att.cell(row=1, column=col)
    cell.font = openpyxl.styles.Font(bold=True, color='FFFFFF')
    cell.fill = openpyxl.styles.PatternFill(start_color='1E3A5F', end_color='1E3A5F', fill_type='solid')
    cell.alignment = openpyxl.styles.Alignment(horizontal='center')

DASH = '\u2014'
summary = {}
for r in rows:
    ci = r['check_in']  if r['check_in']  and r['check_in']  != DASH else DASH
    co = r['check_out'] if r['check_out'] and r['check_out'] != DASH else DASH
    hrs_str = f"{r['working_hours']:.2f} hrs" if r['working_hours'] else '0.00 hrs'
    ws_att.append([r['date'], r['user_id'], r['name'], ci, co, hrs_str, r['status'], r['location'] or ''])
    d = r['date']
    if d not in summary:
        summary[d] = {'total':0,'present':0,'absent':0,'late':0,'on_leave':0}
    summary[d]['total'] += 1
    st = (r['status'] or '').lower()
    if st in ('present', 'logged in'):
        summary[d]['present'] += 1
    elif st == 'late':
        summary[d]['late'] += 1
        summary[d]['present'] += 1
    elif st == 'on leave':
        summary[d]['on_leave'] += 1
    else:
        summary[d]['absent'] += 1

col_widths = [12,12,22,12,12,15,12,30]
for i, w in enumerate(col_widths, 1):
    ws_att.column_dimensions[openpyxl.utils.get_column_letter(i)].width = w

# Rebuild Summary sheet
if 'Summary' in wb.sheetnames:
    del wb['Summary']
ws_sum = wb.create_sheet('Summary')
sum_headers = ['Date','Total','Present','Absent','Late','On Leave','Attendance %']
ws_sum.append(sum_headers)
for col in range(1, len(sum_headers)+1):
    cell = ws_sum.cell(row=1, column=col)
    cell.font = openpyxl.styles.Font(bold=True, color='FFFFFF')
    cell.fill = openpyxl.styles.PatternFill(start_color='1E3A5F', end_color='1E3A5F', fill_type='solid')
    cell.alignment = openpyxl.styles.Alignment(horizontal='center')
for d in sorted(summary.keys(), reverse=True):
    s = summary[d]
    pct = round((s['present']/s['total'])*100, 1) if s['total'] else 0
    ws_sum.append([d, s['total'], s['present'], s['absent'], s['late'], s['on_leave'], f'{pct}%'])

wb.save(EXCEL_FILE)
print(f'Done! Rows written to Attendance sheet: {ws_att.max_row - 1}')
print(f'Dates in Summary: {sorted(summary.keys(), reverse=True)}')
